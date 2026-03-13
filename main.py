"""CLI entry point for the radiology residency schedule maker."""

from __future__ import annotations

import random
import sys
from pathlib import Path

import click
import openpyxl

from schedule_maker.io.excel_reader import ExcelReader
from schedule_maker.io.excel_writer import ExcelWriter
from schedule_maker.io.prefs_parser import PrefsParser
from schedule_maker.models.resident import Resident
from schedule_maker.models.schedule import ScheduleGrid, compute_blocks
from schedule_maker.models.constraints import NFRules
from schedule_maker.phases.r1_assignment import assign_r1_tracks
from schedule_maker.phases.r2_assignment import assign_r2_tracks, print_r2_assignment_matrix
from schedule_maker.phases.r3_builder import assign_r3_fixed, fill_r3_clinical
from schedule_maker.phases.r4_builder import assign_r4_fixed, fill_r4_clinical
from schedule_maker.phases.night_float import assign_night_float
from schedule_maker.phases.sampler import resolve_samplers
from schedule_maker.validation.report import (
    generate_report, generate_preference_report, generate_satisfaction_report,
    compute_r3r4_satisfaction,
)
from schedule_maker.io.dryrun_writer import write_dryrun_xlsx
from schedule_maker.io.prefs_writer import write_preferences


def apply_comment_overrides(residents: list[Resident]) -> None:
    """Apply soft constraints derived from resident free-text survey comments.

    Adds extra no-call dates, section preference boosts, NF timing
    preferences, schedule weight hints, and block requests.
    """
    from datetime import date

    name_map = {r.name: r for r in residents}

    # ── Extra no-call dates ──
    _extra_nocall: dict[str, list[str]] = {
        "Moushey, Alexander": ["7/10", "7/11"],
        "Talebi, Ramtin": ["8/14", "8/15", "9/3", "9/4", "9/5",
                            "10/8", "10/9", "10/10", "10/11", "10/12"],
        "Liu, Liang Yen": ["9/12", "9/13", "9/14", "9/15", "9/16", "9/17", "9/18", "9/19",
                            "12/18", "12/19", "12/20"],
        "Garigipati, Priya": ["12/14", "12/15", "12/16", "12/17", "12/18"],
        "Lue, Brian": ["12/14", "12/15", "12/16", "12/17", "12/18"],
        "Marsh, Kathleen": ["12/25"],
    }
    for name, dates in _extra_nocall.items():
        res = name_map.get(name)
        if res:
            res.no_call.raw_dates.extend(dates)

    # ── Section preference boosts ──
    _section_boosts: dict[str, dict[str, int]] = {
        "Boddu, Priyanka": {"Smr": 2},
        "Yu, Theresa": {"Smr": 1, "Pcbi": 1, "Mai": 1, "Mb": 1},
        "Hu, Anne": {"Smr": 1, "Mch": 1, "Mb": 1},
        "Bermudez Garolera, Daniela": {"Pcbi": 1, "Mnuc": 1},
        "Deshpande, Nikita": {"Mai": 1, "Smr": 1, "Mb": 1, "Ser": 1, "Peds": 1},
        "Chan, Kaelan": {"Smr": 1, "Mb": 1},
        "Ding, Kevin": {"Mucic": 1, "Smr": 1},
    }
    for name, boosts in _section_boosts.items():
        res = name_map.get(name)
        if not res:
            continue
        if res.section_prefs is None:
            from schedule_maker.models.resident import SectionPrefs
            res.section_prefs = SectionPrefs()
        for code, boost in boosts.items():
            res.section_prefs.scores[code] = res.section_prefs.scores.get(code, 0) + boost

    # ── NF timing preferences ──
    _nf_timing: dict[str, str] = {
        "Moushey, Alexander": "avoid-july",
        "Chan, Shin Mei": "early-holidays-ok",
        "Marsh, Kathleen": "late-fall",
        "Diwanji, Devan": "avoid-core-adjacent",
        "Rincon-Hekking, John": "late",
        "Talebi, Ramtin": "holidays-ok",
    }
    for name, pref in _nf_timing.items():
        res = name_map.get(name)
        if res:
            res.nf_timing_pref = pref

    # ── Schedule weight ──
    _schedule_weight: dict[str, str] = {
        "Chan, Shin Mei": "front-heavy",
        "Garigipati, Priya": "front-heavy",
        "Liu, Liang Yen": "front-heavy",
        "Marsh, Kathleen": "back-heavy",
        "Rincon-Hekking, John": "back-heavy",
    }
    for name, weight in _schedule_weight.items():
        res = name_map.get(name)
        if res:
            res.schedule_weight = weight

    # ── Block requests ──
    _block_requests: dict[str, dict[int, str]] = {
        "Stahl, Maximilian": {2: "CEP"},
        "Deshpande, Nikita": {9: "CEP"},
    }
    for name, requests in _block_requests.items():
        res = name_map.get(name)
        if res:
            res.block_requests = requests


@click.group()
def cli():
    """Radiology residency schedule maker."""
    pass


@cli.command()
@click.argument("schedule_file", type=click.Path(exists=True))
@click.argument("prefs_file", type=click.Path(exists=True), required=False, default=None)
@click.option("--output", "-o", type=click.Path(), default=None,
              help="Output .xlsm file path (default: <input>_output.xlsm)")
@click.option("--year", "-y", type=int, default=None,
              help="Academic year start (e.g. 2025). Auto-detected if not specified.")
@click.option("--dry-run", is_flag=True, help="Validate only, don't write to Excel")
@click.option("--core-block", type=int, default=13,
              help="Block number for CORE exam (default: 13)")
@click.option("--lc-weeks", type=str, default=None,
              help="Comma-separated week numbers for LC (e.g. 46,47,48,49)")
@click.option("--core-weeks", type=str, default=None,
              help="Comma-separated week numbers for CORE (e.g. 50,51)")
@click.option("--num-trials", default=1, show_default=True,
              help="Run N stochastic fill trials and keep the best by resident satisfaction (1=deterministic).")
@click.option("--seed", default=None, type=int,
              help="Random seed for reproducibility across trials.")
@click.option("--shuffle-residents/--no-shuffle-residents", default=False,
              help="Randomize resident processing order within pathway groups.")
@click.option("--shuffle-blocks/--no-shuffle-blocks", default=False,
              help="Randomize block processing order within schedule weight groups.")
@click.option("--top-k-sample", default=1, type=int, show_default=True,
              help="Sample from top K rotations instead of best (1=deterministic).")
@click.option("--search-mode", type=click.Choice(["none", "sampling", "hybrid"]),
              default="none", show_default=True,
              help="Optimization mode: none=deterministic, sampling=MC trials, hybrid=MC+local search.")
@click.option("--refine-top", default=3, type=int, show_default=True,
              help="Number of top MC trials to refine with local search (hybrid mode only).")
@click.option("--local-search-iters", default=500, type=int, show_default=True,
              help="Number of local search iterations per candidate (hybrid mode only).")
def build(schedule_file: str, prefs_file: str | None, output: str | None,
          year: int | None, dry_run: bool, core_block: int,
          lc_weeks: str | None, core_weeks: str | None,
          num_trials: int, seed: int | None,
          shuffle_residents: bool, shuffle_blocks: bool, top_k_sample: int,
          search_mode: str, refine_top: int, local_search_iters: int):
    """Build the full schedule from a .xlsm template and optional preferences file.

    If PREFS_FILE is omitted, preferences are read from the Preferences tab
    of the schedule file (populated via 'import-prefs' command).
    """

    schedule_path = Path(schedule_file)
    prefs_path = Path(prefs_file) if prefs_file else None

    click.echo(f"Reading schedule template: {schedule_path.name}")
    if prefs_path:
        click.echo(f"Reading preferences: {prefs_path.name}")
    else:
        click.echo("Reading preferences from Preferences tab")

    # ── Phase 0: Load data ────────────────────────────────────
    with ExcelReader(schedule_path) as reader:
        if year is None:
            year = reader.read_academic_year()
        click.echo(f"Academic year: {year}-{year + 1}")

        rotation_codes = reader.read_rotation_codes()
        click.echo(f"Loaded {len(rotation_codes)} rotation codes")

        residents = reader.read_roster()
        click.echo(f"Loaded {len(residents)} residents")

        # Load historical data
        reader.read_historical_assignments(residents)

        # Load tracks
        r1_tracks = reader.read_r1_tracks()
        r2_tracks = reader.read_r2_tracks()
        click.echo(f"Loaded {len(r1_tracks)} R1 tracks, {len(r2_tracks)} R2 tracks")

        # Load schedule structure
        # (diagnostics about track count vs class size printed after residents are split)
        base_structure = reader.read_base_schedule_structure()
        nf_structure = reader.read_night_float_structure()

        # Load dynamic staffing constraints from Base Schedule rows 101-151
        staffing_constraints = reader.read_staffing_constraints()
        click.echo(f"Loaded {len(staffing_constraints)} staffing constraints")

        # Load NF Recs for dynamic NF rules
        nf_recs = reader.read_nf_recs()

    # Load preferences (before R3-4 Recs so authoritative pathways override)
    if prefs_path:
        click.echo("Parsing preference responses...")
        with PrefsParser(prefs_path) as parser:
            parser.parse_all(residents, academic_year=year)
    else:
        click.echo("Reading preferences from Preferences tab...")
        with ExcelReader(schedule_path) as reader:
            reader.read_preferences_tab(residents)

    # Load R3-4 recommendations AFTER preferences — R3-4 Recs is the
    # authoritative source for pathway flags, overriding self-reported prefs
    if prefs_path:
        # Single-step mode: read only static data (pathway flags + FSE) from
        # R3-4 Recs, then compute recommended_blocks in Python (avoids needing
        # Excel to recalculate formula-driven columns)
        with ExcelReader(schedule_path) as reader:
            reader.read_r34_recs_static(residents)
        from schedule_maker.validation.graduation import compute_r34_recs, compute_intraclass_deficiencies
        intraclass_defs = compute_intraclass_deficiencies(residents)
        compute_r34_recs(residents, intraclass_deficiencies=intraclass_defs)
        click.echo(f"Computed R3-4 recommended blocks (intraclass deficiencies: {len(intraclass_defs)} R3s)")
    else:
        # Two-step mode: R3-4 Recs formulas already recalculated in Excel
        with ExcelReader(schedule_path) as reader:
            reader.read_r34_recs(residents)

    # ── CEP/Research approvals from cep.md ─────────────────────
    from schedule_maker.io.cep_parser import parse_cep_rules
    cep_path = Path(__file__).resolve().parent / ".claude" / "rules" / "cep.md"
    if cep_path.exists():
        applied = parse_cep_rules(cep_path, residents)
        for name, counts in applied.items():
            parts = []
            if counts["cep"]:
                parts.append(f"{counts['cep']} CEP")
            if counts["research"]:
                parts.append(f"{counts['research']} Research")
            click.echo(f"  {name}: {', '.join(parts)}")
        click.echo(f"Applied CEP/Research approvals from cep.md ({len(applied)} residents)")
    else:
        click.echo("Warning: cep.md not found — using form-derived CEP/Research values")

    # ── Comment-derived overrides ──────────────────────────────
    apply_comment_overrides(residents)

    # ── Initialize schedule grid ──────────────────────────────
    blocks = compute_blocks(year)
    grid = ScheduleGrid(blocks=blocks)

    r1s = [r for r in residents if r.r_year == 1]
    r2s = [r for r in residents if r.r_year == 2]
    r3s = [r for r in residents if r.r_year == 3]
    r4s = [r for r in residents if r.r_year == 4]
    click.echo(f"R1: {len(r1s)}, R2: {len(r2s)}, R3: {len(r3s)}, R4: {len(r4s)}")

    if len(r1_tracks) < len(r1s):
        click.echo(f"  Warning: Only {len(r1_tracks)} unique R1 tracks for {len(r1s)} residents — "
                   f"{len(r1s) - len(r1_tracks)} will get duplicate tracks")
    if len(r2_tracks) < len(r2s):
        click.echo(f"  Warning: Only {len(r2_tracks)} unique R2 tracks for {len(r2s)} residents — "
                   f"{len(r2s) - len(r2_tracks)} will get duplicate tracks")

    # ── Phase 1: R1 Tracks ────────────────────────────────────
    click.echo("\n--- Phase 1: R1 Track Assignment ---")
    r1_assignments = assign_r1_tracks(r1s, r1_tracks, grid)
    click.echo(f"Assigned {len(r1_assignments)} R1s to tracks")

    # ── Phase 2: R2 Tracks ────────────────────────────────────
    click.echo("\n--- Phase 2: R2 Track Assignment ---")
    click.echo(print_r2_assignment_matrix(r2s, len(r2_tracks)))
    r2_result = assign_r2_tracks(r2s, r2_tracks, grid)
    if r2_result.feasible:
        click.echo(f"Assigned {len(r2_result.assignments)} R2s (total penalty: {r2_result.total_rank_penalty})")
        for name, info in sorted(r2_result.per_resident.items()):
            click.echo(f"  {name}: Track {info['track']} (rank #{info['rank']})")
    else:
        click.echo(f"R2 assignment FAILED: {r2_result.status}")

    # ── Parse week-level LC/CORE options ─────────────────────
    lc_week_list = [int(w.strip()) for w in lc_weeks.split(",")] if lc_weeks else None
    core_week_list = [int(w.strip()) for w in core_weeks.split(",")] if core_weeks else None
    if core_week_list is None and core_block:
        start = (core_block - 1) * 4 + 1
        core_week_list = [start, start + 1]

    # ── Phase 3: R3 Fixed (AIRP + LC + CORE) ──────────────────
    if lc_week_list or core_week_list:
        click.echo(f"\n--- Phase 3: R3 Fixed Assignments (AIRP + LC weeks={lc_week_list}, CORE weeks={core_week_list}) ---")
    else:
        click.echo(f"\n--- Phase 3: R3 Fixed Assignments (AIRP + LC, CORE=block {core_block}) ---")
    r3_fixed = assign_r3_fixed(r3s, grid, core_exam_block=core_block,
                               lc_weeks=lc_week_list, core_weeks=core_week_list)
    airp_assignments = r3_fixed["airp_assignments"]
    for name, session in sorted(airp_assignments.items()):
        click.echo(f"  {name}: AIRP session={session}")

    # ── Phase 4: R4 Fixed (commitments) ───────────────────────
    click.echo("\n--- Phase 4: R4 Fixed Commitments ---")
    # Derive T32 clinical blocks from the blocks that contain LC/CORE weeks.
    # Always include the LC block (core_block - 1) so T32s get enough
    # clinical weeks for NF coverage.
    lc_block = core_block - 1
    if lc_week_list or core_week_list:
        all_fixed_weeks = (lc_week_list or []) + (core_week_list or [])
        t32_clinical_blocks = sorted({(w - 1) // 4 + 1 for w in all_fixed_weeks} | {lc_block})
    else:
        t32_clinical_blocks = [lc_block, core_block]
    r4_fixed_meta = assign_r4_fixed(r4s, grid, t32_clinical_blocks=t32_clinical_blocks, lc_block=lc_block)
    for name, meta in sorted(r4_fixed_meta.items()):
        if meta.get("t32_clinical_filled") is not None:
            research = meta.get("research_blocks", 0)
            clinical = meta.get("t32_clinical_filled", {})
            click.echo(f"  {name} [T32]: research={research}, clinical={len(clinical)} "
                        f"({', '.join(f'B{b}={c}' for b, c in sorted(clinical.items()))})")
        else:
            fixed = (meta.get("nrdr_mnuc_blocks", 0) + meta.get("esnr_neuro_blocks", 0)
                     + meta.get("esir_mir_blocks", 0) + meta.get("research_blocks", 0)
                     + meta.get("fse_blocks", 0))
            avail = len(meta.get("available_after_fixed", []))
            click.echo(f"  {name}: fixed={fixed}, available={avail}")

    # ── Phase 5: Night Float ──────────────────────────────────
    click.echo("\n--- Phase 5: Night Float Assignment ---")
    # Use NF Recs tab data when available (two-step mode); otherwise defaults
    # are fine — NFRules() matches the standard rules from goals.md
    nf_rules = NFRules.from_nf_recs(nf_recs) if nf_recs else NFRules()
    click.echo(f"NF rules: R2 census={nf_rules.r2_mnf_census}, "
               f"R3 Mnf census={nf_rules.r3_mnf_census}, R3 Snf2 census={nf_rules.r3_snf2_census}, "
               f"R4 Snf2 census={nf_rules.r4_snf2_census}")
    nf_result = assign_night_float(
        residents=residents, grid=grid,
        rules=nf_rules,
        airp_assignments=airp_assignments,
        staffing_constraints=staffing_constraints or None,
    )
    if nf_result.feasible:
        total_nf = sum(len(v) for v in nf_result.assignments.values())
        click.echo(f"Assigned {total_nf} NF weeks across {len(nf_result.assignments)} residents")
    else:
        click.echo(f"NF assignment FAILED: {nf_result.status}")

    # ── Snapshot state after Phase 5 (before stochastic fill) ─
    _grid_assignments_snap = dict(grid.assignments)
    _schedules_snap = {r.name: dict(r.schedule) for r in residents}

    best_trial_score = float("-inf")
    best_trial_state: tuple | None = None
    base_rng = random.Random(seed)

    # For hybrid mode, collect top K trials for local search refinement
    all_trial_results: list[tuple[float, tuple]] = []

    # Determine effective settings based on search_mode
    effective_shuffle_residents = shuffle_residents
    effective_shuffle_blocks = shuffle_blocks
    effective_top_k_sample = top_k_sample
    effective_num_trials = num_trials

    if search_mode == "sampling":
        # Enable stochastic features if not explicitly set
        if not shuffle_residents:
            effective_shuffle_residents = True
        if not shuffle_blocks:
            effective_shuffle_blocks = True
        if top_k_sample == 1:
            effective_top_k_sample = 3
        if num_trials == 1:
            effective_num_trials = 20
    elif search_mode == "hybrid":
        # Hybrid mode: enable all stochastic features
        if not shuffle_residents:
            effective_shuffle_residents = True
        if not shuffle_blocks:
            effective_shuffle_blocks = True
        if top_k_sample == 1:
            effective_top_k_sample = 3
        if num_trials == 1:
            effective_num_trials = 20

    if effective_num_trials > 1:
        click.echo(f"\nRunning {effective_num_trials} stochastic fill trials (seed={seed}, mode={search_mode})...")

    for trial_idx in range(effective_num_trials):
        # Restore snapshot on all trials after the first
        if trial_idx > 0:
            grid.assignments.clear()
            grid.assignments.update(_grid_assignments_snap)
            for r in residents:
                r.schedule.clear()
                r.schedule.update(_schedules_snap[r.name])

        # Per-trial RNG (None when effective_num_trials=1 → deterministic behavior)
        trial_rng: random.Random | None = None
        if effective_num_trials > 1:
            trial_seed = base_rng.randint(0, 2**31)
            trial_rng = random.Random(trial_seed)

        # ── Phase 6: R3 Clinical Fill ─────────────────────────────
        if effective_num_trials == 1:
            click.echo("\n--- Phase 6: R3 Clinical Fill ---")
        r3_clinical_meta = fill_r3_clinical(
            r3s, grid, staffing_constraints=staffing_constraints or None, rng=trial_rng,
            shuffle_residents=effective_shuffle_residents, shuffle_blocks=effective_shuffle_blocks,
            top_k_sample=effective_top_k_sample)
        # Merge airp info into r3 metadata
        r3_meta_trial: dict = {}
        for res in r3s:
            meta = r3_clinical_meta.get(res.name, {"filled_blocks": {}})
            meta["airp_session"] = airp_assignments.get(res.name, "")
            r3_meta_trial[res.name] = meta
        if effective_num_trials == 1:
            for name, meta in sorted(r3_meta_trial.items()):
                click.echo(f"  {name}: AIRP={meta.get('airp_session', '?')}, "
                            f"filled={len(meta.get('filled_blocks', {}))}")

        # ── Phase 7: R4 Clinical Fill ─────────────────────────────
        if effective_num_trials == 1:
            click.echo("\n--- Phase 7: R4 Clinical Fill ---")
        r4_meta_trial = fill_r4_clinical(
            r4s, grid, all_residents=residents, fixed_meta=r4_fixed_meta,
            staffing_constraints=staffing_constraints or None,
            lc_block=lc_block, rng=trial_rng,
            shuffle_residents=effective_shuffle_residents, shuffle_blocks=effective_shuffle_blocks,
            top_k_sample=effective_top_k_sample,
        )
        if effective_num_trials == 1:
            for name, meta in sorted(r4_meta_trial.items()):
                if meta.get("t32_clinical_filled") is not None:
                    research = meta.get("research_blocks", 0)
                    clinical = meta.get("t32_clinical_filled", {})
                    click.echo(f"  {name} [T32]: research={research}, clinical={len(clinical)} "
                                f"({', '.join(f'B{b}={c}' for b, c in sorted(clinical.items()))})")
                else:
                    fixed = (meta.get("nrdr_mnuc_blocks", 0) + meta.get("esnr_neuro_blocks", 0)
                             + meta.get("esir_mir_blocks", 0) + meta.get("research_blocks", 0)
                             + meta.get("fse_blocks", 0))
                    grad_req = len(meta.get("grad_req_filled", {}))
                    remaining = len(meta.get("remaining_filled", {}))
                    click.echo(f"  {name}: fixed={fixed}, grad_req={grad_req}, fill={remaining}")

        # ── Phase 8: Sampler Resolution ───────────────────────────
        if effective_num_trials == 1:
            click.echo("\n--- Phase 8: Sampler Resolution ---")
        sampler_replacements_trial = resolve_samplers(
            r1s, grid, all_residents=residents,
            staffing_constraints=staffing_constraints or None,
        )
        if effective_num_trials == 1:
            total_replaced = sum(len(v) for v in sampler_replacements_trial.values())
            click.echo(f"Replaced {total_replaced} Msamp weeks across {len(sampler_replacements_trial)} R1s")

        # ── Score this trial ───────────────────────────────────────
        if effective_num_trials > 1:
            trial_score = compute_r3r4_satisfaction(
                residents, grid, nf_result=nf_result,
                r3_meta=r3_meta_trial, r4_meta=r4_meta_trial,
                sampler_replacements=sampler_replacements_trial,
            )
            click.echo(f"  Trial {trial_idx + 1}/{effective_num_trials}: satisfaction={trial_score:.1f}")

            trial_state = (
                dict(grid.assignments),
                {r.name: dict(r.schedule) for r in residents},
                r3_meta_trial,
                r4_meta_trial,
                sampler_replacements_trial,
            )

            # Collect all trials for hybrid mode
            all_trial_results.append((trial_score, trial_state))

            if trial_score > best_trial_score:
                best_trial_score = trial_score
                best_trial_state = trial_state

    # ── Hybrid mode: Local search refinement on top K trials ───
    if search_mode == "hybrid" and effective_num_trials > 1 and all_trial_results:
        from schedule_maker.optimization import local_search_refine, SearchConfig

        # Sort trials by score and take top K
        all_trial_results.sort(key=lambda x: -x[0])
        top_k = min(refine_top, len(all_trial_results))
        top_candidates = all_trial_results[:top_k]

        click.echo(f"\n--- Hybrid Mode: Refining top {top_k} trials with local search ---")
        refined_results: list[tuple[float, tuple]] = []

        for rank, (pre_score, state) in enumerate(top_candidates):
            # Restore state
            grid.assignments.clear()
            grid.assignments.update(state[0])
            for r in residents:
                r.schedule.clear()
                r.schedule.update(state[1].get(r.name, {}))

            # Run local search
            search_config = SearchConfig(iterations=local_search_iters)
            refine_rng = random.Random(base_rng.randint(0, 2**31))
            refined_score, stats = local_search_refine(
                residents, grid, search_config, refine_rng,
                nf_result=nf_result,
                r3_meta=state[2],
                r4_meta=state[3],
                staffing_constraints=staffing_constraints,
            )
            click.echo(f"  Candidate {rank + 1}: {pre_score:.1f} -> {refined_score:.1f} "
                       f"(+{stats['improvement']:.1f}, {stats['accepted']}/{stats['iterations']} accepted)")

            # Capture refined state
            refined_state = (
                dict(grid.assignments),
                {r.name: dict(r.schedule) for r in residents},
                state[2],  # r3_meta unchanged
                state[3],  # r4_meta unchanged
                state[4],  # sampler_replacements unchanged
            )
            refined_results.append((refined_score, refined_state))

        # Select best refined trial
        refined_results.sort(key=lambda x: -x[0])
        best_trial_score = refined_results[0][0]
        best_trial_state = refined_results[0][1]
        click.echo(f"  Best refined score: {best_trial_score:.1f}")

    # ── Restore best trial state (only when effective_num_trials > 1) ───
    if effective_num_trials > 1 and best_trial_state is not None:
        best_assignments, best_schedules, r3_meta_trial, r4_meta_trial, sampler_replacements_trial = best_trial_state
        grid.assignments.clear()
        grid.assignments.update(best_assignments)
        for r in residents:
            r.schedule.clear()
            r.schedule.update(best_schedules[r.name])
        click.echo(f"Selected trial with satisfaction={best_trial_score:.1f}")

        # Print per-resident summary for best trial
        click.echo("\n--- Phase 6: R3 Clinical Fill (best trial) ---")
        for name, meta in sorted(r3_meta_trial.items()):
            click.echo(f"  {name}: AIRP={meta.get('airp_session', '?')}, "
                        f"filled={len(meta.get('filled_blocks', {}))}")

        click.echo("\n--- Phase 7: R4 Clinical Fill (best trial) ---")
        for name, meta in sorted(r4_meta_trial.items()):
            if meta.get("t32_clinical_filled") is not None:
                research = meta.get("research_blocks", 0)
                clinical = meta.get("t32_clinical_filled", {})
                click.echo(f"  {name} [T32]: research={research}, clinical={len(clinical)} "
                            f"({', '.join(f'B{b}={c}' for b, c in sorted(clinical.items()))})")
            else:
                fixed = (meta.get("nrdr_mnuc_blocks", 0) + meta.get("esnr_neuro_blocks", 0)
                         + meta.get("esir_mir_blocks", 0) + meta.get("research_blocks", 0)
                         + meta.get("fse_blocks", 0))
                grad_req = len(meta.get("grad_req_filled", {}))
                remaining = len(meta.get("remaining_filled", {}))
                click.echo(f"  {name}: fixed={fixed}, grad_req={grad_req}, fill={remaining}")

        click.echo("\n--- Phase 8: Sampler Resolution (best trial) ---")
        total_replaced = sum(len(v) for v in sampler_replacements_trial.values())
        click.echo(f"Replaced {total_replaced} Msamp weeks across {len(sampler_replacements_trial)} R1s")

    # Expose final trial results under canonical names
    r3_meta = r3_meta_trial
    r4_meta = r4_meta_trial
    sampler_replacements = sampler_replacements_trial

    # ── Phase 9: Validation ───────────────────────────────────
    click.echo("\n--- Phase 9: Validation ---")
    build_metadata = {**r3_meta, **r4_meta}
    report = generate_report(
        residents, grid,
        build_metadata=build_metadata,
        nf_violations=nf_result.violations if nf_result.feasible else None,
        staffing_constraints=staffing_constraints or None,
    )
    click.echo(report)

    # Preference fulfillment report
    pref_report = generate_preference_report(
        residents, grid,
        r2_result=r2_result,
        r3_meta=r3_meta,
        sampler_replacements=sampler_replacements,
        r4_meta=r4_meta,
    )
    click.echo(pref_report)

    # Satisfaction report
    satisfaction = generate_satisfaction_report(
        residents, grid,
        nf_result=nf_result,
        r2_result=r2_result,
        r3_meta=r3_meta,
        r4_meta=r4_meta,
        sampler_replacements=sampler_replacements,
    )
    click.echo(satisfaction)

    # ── Phase 10: Write to Excel ──────────────────────────────
    if not dry_run:
        click.echo(f"\n--- Phase 10: Writing to Excel ---")

        # Single-step mode: also write preferences into the output file
        if prefs_path and output:
            click.echo("Writing preferences to Preferences tab...")
            nocall_datetimes, section_requests = _read_form_extras(prefs_path, residents)
            prefs_output = write_preferences(
                source_path=schedule_path,
                residents=residents,
                output_path=Path(output),
                nocall_datetimes=nocall_datetimes,
                section_requests=section_requests,
            )
            # ExcelWriter copies source→output; use the prefs-populated file
            # as source so it writes schedule on top of the prefs data
            write_source = prefs_output
        else:
            write_source = schedule_path

        with ExcelWriter(write_source, output) as writer:
            writer.set_academic_year(year)

            # Build assignment dicts
            all_assignments = {}
            for res in residents:
                if res.schedule:
                    all_assignments[res.name] = dict(res.schedule)

            nf_assignments = {}
            for name, nf_list in nf_result.assignments.items():
                nf_assignments[name] = {w: code for w, code in nf_list}

            track_map = {r.name: r.track_number for r in residents
                         if r.track_number is not None}

            writer.write_base_schedule(
                all_assignments,
                base_structure["resident_rows"],
                track_map=track_map,
            )

            # NF writer
            writer.write_night_float(
                nf_assignments,
                nf_structure["resident_rows"],
            )

        click.echo(f"Schedule written to: {output or schedule_path.stem + '_output.xlsm'}")
    else:
        dryrun_path = write_dryrun_xlsx(
            "dryrun_schedule.xlsx", residents, grid,
            staffing_constraints=staffing_constraints or None,
        )
        click.echo(f"\nDry-run schedule written to: {dryrun_path}")


@cli.command()
@click.argument("schedule_file", type=click.Path(exists=True))
def validate(schedule_file: str):
    """Validate an existing schedule file."""
    with ExcelReader(schedule_file) as reader:
        residents = reader.read_roster()
        reader.read_historical_assignments(residents)
        reader.read_r34_recs(residents)

    grid = ScheduleGrid()
    for res in residents:
        for w, code in res.schedule.items():
            grid.assign(res.name, w, code)

    report = generate_report(residents, grid)
    click.echo(report)


@cli.command("import-prefs")
@click.argument("schedule_file", type=click.Path(exists=True))
@click.argument("prefs_file", type=click.Path(exists=True))
@click.option("--output", "-o", type=click.Path(), default=None,
              help="Output .xlsm file path (default: <input>_with_prefs.xlsm)")
def import_prefs(schedule_file: str, prefs_file: str, output: str | None):
    """Import Google Forms responses into the Preferences tab."""
    schedule_path = Path(schedule_file)
    prefs_path = Path(prefs_file)
    output_path = Path(output) if output else None

    click.echo(f"Reading roster from: {schedule_path.name}")
    with ExcelReader(schedule_path) as reader:
        residents = reader.read_roster()
        year = reader.read_academic_year()
    click.echo(f"Loaded {len(residents)} residents")

    # Parse form responses
    click.echo(f"Parsing preferences from: {prefs_path.name}")
    with PrefsParser(prefs_path) as parser:
        parser.parse_all(residents, academic_year=year)

    # Read raw form data for no-call weekend expansion and section requests
    click.echo("Reading raw form data for no-call weekends and section requests...")
    nocall_datetimes, section_requests = _read_form_extras(prefs_path, residents)

    # Write to Preferences tab
    click.echo("Writing preferences to Preferences tab...")
    result_path = write_preferences(
        source_path=schedule_path,
        residents=residents,
        output_path=output_path,
        nocall_datetimes=nocall_datetimes,
        section_requests=section_requests,
    )
    click.echo(f"Output written to: {result_path}")


def _read_form_extras(
    prefs_path: Path, residents: list[Resident],
) -> tuple[dict[str, list], dict[str, str]]:
    """Read raw no-call weekend datetimes and section request text from the form.

    Returns:
        (nocall_datetimes, section_requests) keyed by resident name.
    """
    from datetime import datetime

    wb = openpyxl.load_workbook(str(prefs_path), read_only=True, data_only=True)

    nocall_datetimes: dict[str, list] = {}
    section_requests: dict[str, str] = {}

    # Use shared long-form → Key-code mapping from prefs_writer
    from schedule_maker.io.prefs_writer import _LONG_TO_KEY

    # Detect format: combined vs per-class
    if "Form Responses 1" not in wb.sheetnames:
        wb.close()
        return nocall_datetimes, section_requests

    ws = wb["Form Responses 1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return nocall_datetimes, section_requests

    # PGY mapping for section request column selection
    pgy_to_ryear = {
        "PGY-1": 1, "PGY-2 (R1)": 2, "PGY-3 (R2)": 3, "PGY-4 (R3)": 4,
    }
    # Section request form columns (0-based): R2=10, R3=27, R4=61
    section_req_cols = {2: 10, 3: 27, 4: 61}

    # Build name lookup
    name_map: dict[str, Resident] = {}
    for r in residents:
        name_map[f"{r.first_name} {r.last_name}".lower()] = r
        name_map[r.name.lower()] = r
        name_map[f"{r.last_name}, {r.first_name}".lower()] = r

    # Deduplicate: keep latest per (first, last)
    latest: dict[tuple[str, str], tuple] = {}
    for row in rows[1:]:
        if not any(row):
            continue
        first = str(row[1]).strip() if row[1] else ""
        last = str(row[2]).strip() if row[2] else ""
        if not first or not last:
            continue
        pgy_str = str(row[3]).strip() if row[3] else ""
        if pgy_str not in pgy_to_ryear:
            continue
        key = (first.lower(), last.lower())
        ts = row[0]
        if key not in latest or (ts and (latest[key][0] is None or ts > latest[key][0])):
            latest[key] = (ts, row, pgy_str)

    headers = rows[0]

    for (_first_l, _last_l), (_ts, row, pgy_str) in latest.items():
        first = str(row[1]).strip()
        last = str(row[2]).strip()
        r_year = pgy_to_ryear[pgy_str]

        # Find matching resident
        res = None
        for key_try in (
            f"{first} {last}".lower(),
            f"{last}, {first}".lower(),
        ):
            if key_try in name_map:
                res = name_map[key_try]
                break
        if res is None:
            last_lower = last.lower()
            for r in name_map.values():
                if r.last_name.lower() == last_lower:
                    res = r
                    break
        if res is None:
            continue

        # No-call weekend datetimes (cols 70-71, 0-based)
        dts = []
        for col in (70, 71):
            val = row[col] if col < len(row) else None
            if val is not None and isinstance(val, datetime):
                dts.append(val)
        if dts:
            nocall_datetimes[res.name] = dts

        # AIRP group (R3 only) — detect column dynamically from headers,
        # fall back to col 51 if not found
        if r_year == 3:
            airp_gcol = 51  # default fallback
            for ci in range(50, min(62, len(headers))):
                if headers[ci] is None:
                    continue
                h = str(headers[ci]).lower()
                if "list up to three" in h or ("group" in h and "airp" in h):
                    airp_gcol = ci
                    break
            airp_group = str(row[airp_gcol]).strip() if airp_gcol < len(row) and row[airp_gcol] else ""
            if airp_group:
                if res.airp_prefs is None:
                    from schedule_maker.models.resident import AIRPPrefs
                    res.airp_prefs = AIRPPrefs()
                res.airp_prefs.group_requests = [g.strip() for g in airp_group.split(",") if g.strip()]

        # Pathway org (col 60, R4 only)
        if r_year == 4:
            pathway_org = str(row[60]).strip() if 60 < len(row) and row[60] else ""
            if pathway_org:
                if res.fse_prefs is None:
                    from schedule_maker.models.resident import FSEPrefs
                    res.fse_prefs = FSEPrefs()
                res.fse_prefs.organization = pathway_org

        # Holiday history (cols 68-70)
        holiday_history = []
        for col in (68, 69, 70):
            val = row[col] if col < len(row) else None
            holiday_history.append(str(val).strip() if val is not None else "")
        if any(holiday_history):
            res.no_call.holiday_history = holiday_history

        # Holiday work pref (col 73)
        holiday_pref = str(row[73]).strip() if 73 < len(row) and row[73] else ""
        if holiday_pref:
            res.no_call.holiday_work_pref = holiday_pref

        # Section request (class-appropriate column)
        sec_col = section_req_cols.get(r_year)
        if sec_col is not None and sec_col < len(row) and row[sec_col]:
            text = str(row[sec_col]).strip()
            if text:
                # Abbreviate long-form names (R4 col 61 uses full names)
                raw_parts = [p.strip() for p in text.split(",") if p.strip()]
                parts: list[str] = []
                for p in raw_parts:
                    mapped = _LONG_TO_KEY.get(p)
                    if mapped:
                        parts.extend(mapped)
                    else:
                        parts.append(p)
                section_requests[res.name] = ",".join(parts)

    wb.close()
    return nocall_datetimes, section_requests


@cli.command()
@click.argument("schedule_file", type=click.Path(exists=True))
@click.option("--output", "-o", type=click.Path(), default=None,
              help="Write report to file instead of stdout")
def stats(schedule_file: str, output: str | None):
    """Generate an anonymized equity report for resident distribution."""
    from schedule_maker.validation.equity_report import generate_equity_report

    schedule_path = Path(schedule_file)
    click.echo(f"Reading schedule: {schedule_path.name}")

    with ExcelReader(schedule_path) as reader:
        year = reader.read_academic_year()
        residents = reader.read_roster()
        click.echo(f"Loaded {len(residents)} residents (AY {year}-{year + 1})")

        reader.read_historical_assignments(residents)
        reader.read_r34_recs(residents)

        staffing_constraints = reader.read_staffing_constraints()

        # Try reading preferences for R2 track data
        try:
            reader.read_preferences_tab(residents)
        except Exception:
            pass  # Preferences tab may not exist or be populated

        nf_data = reader.read_schedule_assignments(residents)

    # Reconstruct ScheduleGrid
    blocks = compute_blocks(year)
    grid = ScheduleGrid(blocks=blocks)
    for res in residents:
        for w, code in res.schedule.items():
            grid.assign(res.name, w, code)
    for name, nf_weeks in nf_data.items():
        for w, code in nf_weeks.items():
            grid.assign_nf(name, w, code)

    report = generate_equity_report(
        residents, grid,
        staffing_constraints=staffing_constraints or None,
    )

    if output:
        Path(output).write_text(report)
        click.echo(f"Report written to: {output}")
    else:
        click.echo(report)


if __name__ == "__main__":
    cli()
