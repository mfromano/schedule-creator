"""Dry-run spreadsheet output — standalone .xlsx with schedule, staffing, and graduation data."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from schedule_maker.models.resident import Resident, Pathway
from schedule_maker.models.schedule import ScheduleGrid
from schedule_maker.models.rotation import (
    get_hospital_system, HospitalSystem, is_night_float,
    NM_PARTIAL_CREDIT_ROTATIONS, NM_PARTIAL_RATIO, fse_to_base_code,
)
from schedule_maker.models.constraints import StaffingConstraint
from schedule_maker.validation.staffing import ROTATION_MINIMUMS, ROTATION_MAXIMUMS


# ── Fills ──────────────────────────────────────────────────────────
FILL_NF = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")       # light blue
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # red
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # green
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # yellow
FILL_ORANGE = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")    # orange
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")    # header blue
FONT_BOLD = Font(bold=True)
FONT_HEADER = Font(bold=True, size=9)
FONT_SMALL = Font(size=8)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _week_header(week: int) -> str:
    """Block X WY label for a 1-based week number."""
    block = (week - 1) // 4 + 1
    w_in_block = (week - 1) % 4 + 1
    return f"B{block} W{w_in_block}"


def _sorted_residents(residents: list[Resident]) -> list[Resident]:
    """Sort residents by (r_year, last_name)."""
    return sorted(residents, key=lambda r: (r.r_year, r.last_name or r.name))


def _pathway_label(res: Resident) -> str:
    parts = []
    if res.is_esir:
        parts.append("ESIR")
    if res.is_esnr:
        parts.append("ESNR")
    if res.is_t32:
        parts.append("T32")
    if res.is_nrdr:
        parts.append("NRDR")
    return "/".join(parts) if parts else ""


def _no_call_weeks(res: Resident, grid: ScheduleGrid) -> set[int]:
    """Return set of week numbers that overlap with the resident's no-call dates."""
    if not grid.blocks or not res.no_call.raw_dates:
        return set()
    nc_dates = set(res.no_call.raw_dates)
    result = set()
    week_num = 0
    for block in grid.blocks:
        for mon, _fri in block.weeks:
            week_num += 1
            # Check if any day Mon-Sun of this week overlaps a no-call date
            from datetime import timedelta
            for d in range(7):
                if mon + timedelta(days=d) in nc_dates:
                    result.add(week_num)
                    break
    return result


# ── Sheet 1: Schedule ──────────────────────────────────────────────

def _write_schedule_sheet(
    ws, residents: list[Resident], grid: ScheduleGrid, num_weeks: int,
    staffing_constraints: list[StaffingConstraint] | None = None,
):
    sorted_res = _sorted_residents(residents)

    # Header row
    ws.cell(row=1, column=1, value="Resident").font = FONT_BOLD
    ws.cell(row=1, column=2, value="R").font = FONT_BOLD
    track_cell = ws.cell(row=1, column=3, value="Track")
    track_cell.font = FONT_BOLD
    track_cell.fill = FILL_HEADER
    track_cell.alignment = ALIGN_CENTER
    week_offset = 3  # weeks start at column 4
    for w in range(1, num_weeks + 1):
        cell = ws.cell(row=1, column=w + week_offset, value=_week_header(w))
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 6

    # Resident rows
    for i, res in enumerate(sorted_res):
        row = i + 2
        ws.cell(row=row, column=1, value=res.name)
        ws.cell(row=row, column=2, value=f"R{res.r_year}").alignment = ALIGN_CENTER
        if res.r_year in (1, 2) and res.track_number is not None:
            ws.cell(row=row, column=3, value=res.track_number).alignment = ALIGN_CENTER
        for w in range(1, num_weeks + 1):
            code = grid.get(res.name, w)
            cell = ws.cell(row=row, column=w + week_offset, value=code or "")
            cell.font = FONT_SMALL
            cell.alignment = ALIGN_CENTER
            if code and is_night_float(code):
                cell.fill = FILL_NF

    # Blank separator row
    sep_row = len(sorted_res) + 3

    # Staffing rows
    staff_row = sep_row + 1
    ws.cell(row=staff_row - 1, column=1, value="STAFFING").font = FONT_BOLD

    # Use dynamic constraints if available, otherwise fall back to ROTATION_MINIMUMS
    if staffing_constraints:
        min_entries = [(sc.label, sc.rotation_codes, sc.min_count) for sc in staffing_constraints]
    else:
        min_entries = [(label, codes, min_req) for label, (codes, min_req) in ROTATION_MINIMUMS.items()]

    for label, codes, min_req in min_entries:
        ws.cell(row=staff_row, column=1, value=label).font = FONT_BOLD
        ws.cell(row=staff_row, column=2, value=f"≥{min_req}").alignment = ALIGN_CENTER
        for w in range(1, num_weeks + 1):
            assignments = grid.get_week_assignments(w)
            count = sum(1 for c in assignments.values()
                        if c in codes or fse_to_base_code(c) in codes)
            cell = ws.cell(row=staff_row, column=w + week_offset, value=count)
            cell.alignment = ALIGN_CENTER
            if count < min_req:
                cell.fill = FILL_RED
        staff_row += 1

    # Maximum constraint rows
    staff_row += 1  # blank separator
    ws.cell(row=staff_row, column=1, value="MAXIMUMS").font = FONT_BOLD
    staff_row += 1
    for label, (codes, max_allowed) in ROTATION_MAXIMUMS.items():
        ws.cell(row=staff_row, column=1, value=label).font = FONT_BOLD
        ws.cell(row=staff_row, column=2, value=f"≤{max_allowed}").alignment = ALIGN_CENTER
        for w in range(1, num_weeks + 1):
            assignments = grid.get_week_assignments(w)
            count = sum(1 for c in assignments.values()
                        if c in codes or fse_to_base_code(c) in codes)
            cell = ws.cell(row=staff_row, column=w + week_offset, value=count)
            cell.alignment = ALIGN_CENTER
            if count > max_allowed:
                cell.fill = FILL_RED
        staff_row += 1

    # Site summary rows
    staff_row += 1  # blank separator
    ws.cell(row=staff_row, column=1, value="SITE TOTALS").font = FONT_BOLD
    staff_row += 1
    for site_label, system in [("UCSF", HospitalSystem.UCSF),
                                ("ZSFG", HospitalSystem.ZSFG),
                                ("VA", HospitalSystem.VA)]:
        ws.cell(row=staff_row, column=1, value=site_label).font = FONT_BOLD
        for w in range(1, num_weeks + 1):
            assignments = grid.get_week_assignments(w)
            count = sum(1 for c in assignments.values() if get_hospital_system(c) == system)
            cell = ws.cell(row=staff_row, column=w + week_offset, value=count)
            cell.alignment = ALIGN_CENTER
        staff_row += 1


# ── Sheet 2: Graduation ───────────────────────────────────────────

def _write_graduation_sheet(ws, residents: list[Resident]):
    sorted_res = _sorted_residents(residents)

    headers = [
        "Resident", "R", "Pathway",
        "Breast (Req)", "Breast (Total)", "Breast (Deficit)",
        "NucMed (Req)", "NucMed (Total)", "NucMed (Deficit)",
        "ESIR (Req)", "ESIR (Total)", "ESIR (Deficit)",
        "ESNR (Req)", "ESNR (Total)", "ESNR (Deficit)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = FONT_BOLD
        cell.fill = FILL_HEADER
    ws.column_dimensions["A"].width = 22

    breast_codes = {"Pcbi", "Sbi"}
    nm_codes = {"Mnuc", "Vnuc"}
    ir_codes = {"Mir", "Zir", "Sir", "Vir"}
    neuro_codes = {"Mucic", "Smr"}

    for i, res in enumerate(sorted_res):
        row = i + 2
        # Current year weeks
        current: dict[str, float] = {}
        for _w, code in res.schedule.items():
            if code:
                current[code] = current.get(code, 0) + 1

        ws.cell(row=row, column=1, value=res.name)
        ws.cell(row=row, column=2, value=f"R{res.r_year}").alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=_pathway_label(res))

        # Breast (12 weeks)
        breast_req = 12
        breast_total = sum(res.history.get(c, 0) + current.get(c, 0) for c in breast_codes)
        breast_deficit = max(0, breast_req - breast_total)
        ws.cell(row=row, column=4, value=breast_req).alignment = ALIGN_CENTER
        ws.cell(row=row, column=5, value=breast_total).alignment = ALIGN_CENTER
        deficit_cell = ws.cell(row=row, column=6, value=breast_deficit)
        deficit_cell.alignment = ALIGN_CENTER
        deficit_cell.fill = FILL_RED if breast_deficit > 0 else FILL_GREEN

        # NucMed
        nm_total = sum(res.history.get(c, 0) + current.get(c, 0) for c in nm_codes)
        if res.is_nrdr:
            nm_req = 48
            res_weeks = res.history.get("Res", 0) + current.get("Res", 0)
            nm_total = nm_total + res_weeks
        else:
            partial = sum(
                (res.history.get(c, 0) + current.get(c, 0)) * NM_PARTIAL_RATIO
                for c in NM_PARTIAL_CREDIT_ROTATIONS
            )
            nm_total = nm_total + partial
            nm_req = 16
        nm_deficit = max(0, nm_req - nm_total)
        ws.cell(row=row, column=7, value=nm_req).alignment = ALIGN_CENTER
        ws.cell(row=row, column=8, value=round(nm_total, 1)).alignment = ALIGN_CENTER
        deficit_cell = ws.cell(row=row, column=9, value=round(nm_deficit, 1))
        deficit_cell.alignment = ALIGN_CENTER
        deficit_cell.fill = FILL_RED if nm_deficit > 0 else FILL_GREEN

        # ESIR (12 weeks IR) — blank if not applicable
        if res.is_esir:
            esir_req = 12
            ir_total = sum(res.history.get(c, 0) + current.get(c, 0) for c in ir_codes)
            esir_deficit = max(0, esir_req - ir_total)
            ws.cell(row=row, column=10, value=esir_req).alignment = ALIGN_CENTER
            ws.cell(row=row, column=11, value=ir_total).alignment = ALIGN_CENTER
            deficit_cell = ws.cell(row=row, column=12, value=esir_deficit)
            deficit_cell.alignment = ALIGN_CENTER
            deficit_cell.fill = FILL_RED if esir_deficit > 0 else FILL_GREEN

        # ESNR (24 weeks neuro) — blank if not applicable
        if res.is_esnr:
            esnr_req = 24
            neuro_total = sum(res.history.get(c, 0) + current.get(c, 0) for c in neuro_codes)
            esnr_deficit = max(0, esnr_req - neuro_total)
            ws.cell(row=row, column=13, value=esnr_req).alignment = ALIGN_CENTER
            ws.cell(row=row, column=14, value=neuro_total).alignment = ALIGN_CENTER
            deficit_cell = ws.cell(row=row, column=15, value=esnr_deficit)
            deficit_cell.alignment = ALIGN_CENTER
            deficit_cell.fill = FILL_RED if esnr_deficit > 0 else FILL_GREEN


# ── Sheet 3: NF Assignments ───────────────────────────────────────

def _write_nf_sheet(ws, residents: list[Resident], grid: ScheduleGrid, num_weeks: int):
    sorted_res = _sorted_residents(residents)

    # Header
    ws.cell(row=1, column=1, value="Resident").font = FONT_BOLD
    ws.cell(row=1, column=2, value="R").font = FONT_BOLD
    for w in range(1, num_weeks + 1):
        cell = ws.cell(row=1, column=w + 2, value=_week_header(w))
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 4

    for i, res in enumerate(sorted_res):
        row = i + 2
        ws.cell(row=row, column=1, value=res.name)
        ws.cell(row=row, column=2, value=f"R{res.r_year}").alignment = ALIGN_CENTER

        nc_weeks = _no_call_weeks(res, grid)
        for w in range(1, num_weeks + 1):
            nf_code = grid.nf_assignments.get((res.name, w), "")
            cell = ws.cell(row=row, column=w + 2, value=nf_code or "")
            cell.font = FONT_SMALL
            cell.alignment = ALIGN_CENTER
            if nf_code and w in nc_weeks:
                cell.fill = FILL_ORANGE  # NF on no-call week = violation
            elif nf_code:
                cell.fill = FILL_NF
            elif w in nc_weeks:
                cell.fill = FILL_YELLOW  # no-call preference


# ── Public API ─────────────────────────────────────────────────────

def write_dryrun_xlsx(
    path: str | Path,
    residents: list[Resident],
    grid: ScheduleGrid,
    num_weeks: int = 52,
    staffing_constraints: list[StaffingConstraint] | None = None,
) -> Path:
    """Write a dry-run schedule spreadsheet.

    Creates three sheets:
      1. Schedule — full resident × week grid with staffing counts
      2. Graduation — per-resident graduation requirement status
      3. NF Assignments — night float overlay with no-call highlighting

    Returns the path written to.
    """
    path = Path(path)
    wb = Workbook()

    # Sheet 1: Schedule
    ws_sched = wb.active
    ws_sched.title = "Schedule"
    _write_schedule_sheet(ws_sched, residents, grid, num_weeks, staffing_constraints)

    # Sheet 2: Graduation
    ws_grad = wb.create_sheet("Graduation")
    _write_graduation_sheet(ws_grad, residents)

    # Sheet 3: NF Assignments
    ws_nf = wb.create_sheet("NF Assignments")
    _write_nf_sheet(ws_nf, residents, grid, num_weeks)

    wb.save(path)
    return path
