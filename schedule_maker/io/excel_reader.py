"""Read data from the Schedule Creation .xlsm workbook."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from schedule_maker.models.resident import Resident, Pathway
from schedule_maker.models.rotation import RotationCode, SECTION_TO_ROTATION_CODES
from schedule_maker.models.constraints import StaffingConstraint
from schedule_maker.models.schedule import Block


# Rotation aliases: codes that have been renamed/merged.
# When any of these codes are read from Excel, they are mapped to the new code.
_ROTATION_ALIASES: dict[str, str] = {
    "Vn": "Mucic",
}


def _str(val) -> str:
    """Safely convert a cell value to string."""
    if val is None:
        return ""
    return str(val).strip()


def _normalize_rotation(code: str) -> str:
    """Apply rotation aliases (e.g. Vn → Mucic)."""
    return _ROTATION_ALIASES.get(code, code)


def _float(val) -> float:
    """Safely convert a cell value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_pathway(esnr: str, esir: str, t32: str, nrdr: str) -> Pathway:
    """Parse pathway flags from 'x' markers in Historical tab."""
    p = Pathway.NONE
    if _str(esnr).lower() == "x":
        p |= Pathway.ESNR
    if _str(esir).lower() == "x":
        p |= Pathway.ESIR
    if _str(t32).lower() == "x":
        p |= Pathway.T32
    if _str(nrdr).lower() == "x":
        p |= Pathway.NRDR
    return p


@dataclass
class TrackTemplate:
    """A track template: ordered list of rotation codes per biweek."""
    number: int
    label: str  # e.g. "1A", "1B" → simplified to just track number
    # List of (block, biweek, rotation_code) tuples
    assignments: list[tuple[int, str, str]] = field(default_factory=list)

    def to_weekly_schedule(self) -> dict[int, str]:
        """Convert biweekly track to weekly schedule.

        Each block has 4 weeks. Biweek A = weeks 1-2, Biweek B = weeks 3-4.

        "Sx" in the track sheet denotes the Sx/Snf rotation package:
          - Full Sx block (A=Sx, B=Sx): Sx, Snf, Snf, Sx
          - Partial biweek B Sx: Sx, Snf + forward spill Snf, Sx into
            next block (except block 13 end of year)
          - Partial biweek A Sx: standalone Snf, Sx (no backward spill)
          - Block 1 partial biweek A: Snf, Sx (standalone, start of year)
        """
        # Build block composition map
        block_biweeks: dict[int, dict[str, str]] = {}
        for block, biweek, code in self.assignments:
            block_biweeks.setdefault(block, {})[biweek] = code

        full_sx_blocks: set[int] = set()
        for block, bws in block_biweeks.items():
            if bws.get("A") == "Sx" and bws.get("B") == "Sx":
                full_sx_blocks.add(block)

        # Standard expansion first
        schedule: dict[int, str] = {}
        for block, biweek, code in self.assignments:
            base_week = (block - 1) * 4 + 1
            if biweek == "A":
                schedule[base_week] = code
                schedule[base_week + 1] = code
            else:
                schedule[base_week + 2] = code
                schedule[base_week + 3] = code

        # Overlay Sx/Snf patterns
        for block, biweek, code in self.assignments:
            if code != "Sx":
                continue
            base_week = (block - 1) * 4 + 1

            if block in full_sx_blocks:
                if biweek == "A":
                    for i, rot in enumerate(["Sx", "Snf", "Snf", "Sx"]):
                        schedule[base_week + i] = rot
            elif biweek == "B":
                # Partial biweek B: Sx, Snf + forward spill Snf, Sx
                schedule[base_week + 2] = "Sx"
                schedule[base_week + 3] = "Snf"
                if block < 13:
                    next_base = block * 4 + 1
                    schedule[next_base] = "Snf"
                    schedule[next_base + 1] = "Sx"
            else:  # biweek A, not in full block
                # In the Excel, the rotation chief manually replaces this
                # partial-A Sx with a clinical rotation.  Our fallback
                # can't replicate that edit, so substitute biweek B's
                # rotation for both biweeks (skip Sx/Snf entirely).
                bw_b_code = block_biweeks.get(block, {}).get("B", "")
                if bw_b_code and bw_b_code != "Sx":
                    schedule[base_week] = bw_b_code
                    schedule[base_week + 1] = bw_b_code
        return schedule


class ExcelReader:
    """Reads all scheduling data from the Schedule Creation .xlsm file."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self._wb = openpyxl.load_workbook(
            str(self.path), read_only=True, data_only=True, keep_vba=True
        )

    def close(self):
        self._wb.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    # ── Overview ──────────────────────────────────────────────

    def read_academic_year(self) -> int:
        """Read target academic year from Overview tab cell B5."""
        ws = self._wb["Overview"]
        val = ws["B5"].value
        if val is None:
            val = ws["B6"].value  # some versions use B6
        return int(val) if val else 2026

    # ── Key (Rotation Codes) ─────────────────────────────────

    def read_rotation_codes(self) -> list[RotationCode]:
        """Parse the Key tab into RotationCode objects."""
        ws = self._wb["Key"]
        codes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            code_val = _str(row[0]) if row[0] else None
            if not code_val:
                continue
            section = _str(row[1]) if len(row) > 1 else ""
            label = _str(row[2]) if len(row) > 2 else ""

            rc = RotationCode(
                code=code_val,
                section=section,
                label=label,
            )
            # Parse PGY eligibility (columns D-G or similar)
            if len(row) > 3:
                for i, pgy in enumerate([1, 2, 3, 4], start=3):
                    if len(row) > i and _str(row[i]).lower() in ("x", "yes", "true", "1"):
                        rc.pgy_eligible.add(pgy)
                        if pgy == 1:
                            rc.r1_eligible = True
                        elif pgy == 2:
                            rc.r2_eligible = True
                        elif pgy == 3:
                            rc.r3_eligible = True
                        elif pgy == 4:
                            rc.r4_eligible = True
            codes.append(rc)
        return codes

    # ── Historical (Roster + History) ────────────────────────

    def _detect_historical_layout(self) -> dict:
        """Auto-detect the Historical tab column layout.

        Two known layouts:
          Layout A (2025-2026): A=Current PGY, B=Future PGY, C=Resident, D+=history
          Layout B (2026-2027): A=Current PGY, B=Resident, C+=ESNR/ESIR/T32/NRDR, G+=history

        Returns dict with 'pgy_col', 'name_col', 'has_future_pgy', 'history_start_col',
        and pathway column indices.
        """
        ws = self._wb["Historical"]
        headers = []
        for row in ws.iter_rows(min_row=2, max_row=2, max_col=10, values_only=True):
            headers = [_str(v) for v in row]
            break

        if len(headers) > 1 and "future" in headers[1].lower():
            # Layout A: A=Current PGY, B=Future PGY, C=Resident
            return {
                "has_future_pgy": True,
                "pgy_col": 1,      # Future PGY in column B
                "name_col": 2,     # Resident in column C
                "history_start_col": 3,
                "pathway_cols": None,  # No pathway columns in this layout
            }
        else:
            # Layout B: A=Current PGY, B=Resident, C=ESNR, D=ESIR, E=T32, F=NRDR
            return {
                "has_future_pgy": False,
                "pgy_col": 0,      # Current PGY in column A (will increment by 1)
                "name_col": 1,     # Resident in column B
                "history_start_col": 6,  # Block data starts after pathway cols
                "pathway_cols": {"esnr": 2, "esir": 3, "t32": 4, "nrdr": 5},
            }

    def read_roster(self) -> list[Resident]:
        """Parse the Historical tab to get all residents with PGY.

        Supports two layouts:
          Layout A: Has 'Future PGY' column — use directly.
          Layout B: Only 'Current PGY' — increment by 1 to get target year PGY.

        Target year PGY: PGY-2=R1, PGY-3=R2, PGY-4=R3, PGY-5=R4.
        PGY-1 (interns) and PGY-6+ (graduated) are skipped.
        """
        ws = self._wb["Historical"]
        layout = self._detect_historical_layout()
        residents = []

        max_col = max(layout["name_col"] + 1, 7)
        for row in ws.iter_rows(min_row=3, max_col=max_col, values_only=True):
            raw_pgy = row[layout["pgy_col"]]
            name = _str(row[layout["name_col"]])
            if not name or not raw_pgy:
                continue
            try:
                pgy_int = int(raw_pgy)
            except (ValueError, TypeError):
                continue

            # If no Future PGY column, increment current PGY by 1
            if not layout["has_future_pgy"]:
                pgy_int += 1

            r_year = pgy_int - 1  # PGY-2=R1, PGY-3=R2, etc.
            if r_year < 1 or r_year > 4:
                continue  # Skip interns (PGY-1) and graduated (PGY-6+)

            # Parse name ("Last, First" format)
            parts = name.split(",", 1)
            last = parts[0].strip()
            first = parts[1].strip() if len(parts) > 1 else ""

            res = Resident(
                name=name,
                first_name=first,
                last_name=last,
                pgy=pgy_int,
                r_year=r_year,
            )

            # Parse pathway flags if present in this layout
            if layout["pathway_cols"]:
                pw = layout["pathway_cols"]
                res.pathway = _parse_pathway(
                    _str(row[pw["esnr"]]) if len(row) > pw["esnr"] else "",
                    _str(row[pw["esir"]]) if len(row) > pw["esir"] else "",
                    _str(row[pw["t32"]]) if len(row) > pw["t32"] else "",
                    _str(row[pw["nrdr"]]) if len(row) > pw["nrdr"] else "",
                )

            residents.append(res)

        return residents

    def read_historical_assignments(self, residents: list[Resident]) -> None:
        """Read 4 years of weekly rotation history from Historical tab.

        Updates residents' history dict in-place with cumulative weeks per rotation.
        Auto-detects layout to find the correct name and history start columns.
        """
        ws = self._wb["Historical"]
        layout = self._detect_historical_layout()
        name_col = layout["name_col"]
        history_start = layout["history_start_col"]
        name_to_resident = {r.name: r for r in residents}

        for row in ws.iter_rows(min_row=3, values_only=True):
            name = _str(row[name_col]) if len(row) > name_col else ""
            if name not in name_to_resident:
                continue

            res = name_to_resident[name]
            for col_idx in range(history_start, len(row)):
                code = _normalize_rotation(_str(row[col_idx]))
                if code and code not in ("", "0"):
                    res.history[code] = res.history.get(code, 0) + 1

    # ── Historical Tabulation ────────────────────────────────

    def read_historical_tabulation(self, residents: list[Resident]) -> None:
        """Read cumulative rotation weeks from Historical Tabulation tab.

        This is a pre-computed summary that may be more accurate than
        raw weekly counting. Updates resident history in-place.
        """
        ws = self._wb["Historical Tabulation"]

        # Read header row to identify rotation columns
        headers = []
        for cell in ws[1]:
            headers.append(_str(cell.value))

        # Find resident rows and map columns
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = _str(row[1]) if len(row) > 1 else ""
            matching = [r for r in residents if r.name == name]
            if not matching:
                continue
            res = matching[0]

            # Tabulation columns contain section totals
            for i, header in enumerate(headers):
                if i < 2 or not header:
                    continue
                val = _float(row[i]) if len(row) > i else 0
                if val > 0:
                    res.history[header] = val

    # ── Transfers ────────────────────────────────────────────

    def read_transfers(self) -> dict[str, dict[str, float]]:
        """Read transfer credit weeks from Transfers tab.

        Returns {name: {credit_type: weeks}}.
        """
        ws = self._wb["Transfers"]
        transfers = {}
        headers = []
        for row in ws.iter_rows(max_row=1, values_only=True):
            headers = [_str(v) for v in row]
            break

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = _str(row[0])
            if not name:
                continue
            credits = {}
            for i in range(1, len(headers)):
                if len(row) > i and row[i]:
                    credits[headers[i]] = _float(row[i])
            transfers[name] = credits

        return transfers

    # ── R1/R2 Tracks ─────────────────────────────────────────

    def _read_tracks(self, tab_name: str) -> list[TrackTemplate]:
        """Read track templates from R1 Tracks or R2 Tracks tab.

        Preferred: read cached grid values directly from the track columns,
        which reflect any manual edits the rotation chief made.

        Fallback (if grid values are mostly None — file wasn't recalculated
        in Excel): derive tracks from base sequence using the rotation formula
          position = ((track_num - 1) + (block - 1) * 2) % seq_len + 1
        """
        ws = self._wb[tab_name]

        # --- Determine track count and column mapping from header labels ---
        header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=False))[0]
        track_numbers = set()
        # Map (track_num, biweek) → 0-based column index
        track_col_map: dict[tuple[int, str], int] = {}
        for i, cell in enumerate(header_row):
            if i < 6:
                continue
            val = _str(cell.value)
            if not val:
                continue
            if val[-1] in ("A", "B") and val[:-1].isdigit():
                tnum = int(val[:-1])
                bw = val[-1]
                track_numbers.add(tnum)
                track_col_map[(tnum, bw)] = i
            # Stop at the first non-track label (e.g. "MISSING", "Section")
            elif not val.isdigit():
                break

        if not track_numbers:
            return []
        track_count = max(track_numbers)

        # --- Determine how many data rows to read ---
        # Row 7+ contains the grid; each block has 2 rows (A, B), 13 blocks = 26 rows.
        num_blocks = 13
        expected_rows = num_blocks * 2  # 26

        # --- Attempt to read cached grid values from track columns ---
        max_col_needed = max(track_col_map.values()) + 1 if track_col_map else 3
        grid_rows = list(ws.iter_rows(
            min_row=7, max_row=7 + expected_rows - 1,
            max_col=max_col_needed, values_only=True,
        ))

        # Count how many non-None grid cells we find
        grid_non_none = 0
        grid_total = 0
        for row in grid_rows:
            for (_tnum, _bw), col_idx in track_col_map.items():
                if col_idx < len(row):
                    grid_total += 1
                    if row[col_idx] is not None:
                        grid_non_none += 1

        use_grid = grid_total > 0 and grid_non_none >= grid_total * 0.5

        if use_grid:
            # --- Build tracks from cached grid values ---
            tracks = [TrackTemplate(number=i + 1, label=f"Track {i + 1}")
                      for i in range(track_count)]

            for tnum in range(1, track_count + 1):
                for bw in ("A", "B"):
                    col_idx = track_col_map.get((tnum, bw))
                    if col_idx is None:
                        continue
                    for row_idx, row in enumerate(grid_rows):
                        if col_idx >= len(row):
                            continue
                        code = _str(row[col_idx])
                        if not code:
                            continue
                        code = _normalize_rotation(code)
                        # row_idx 0 = block 1 A, 1 = block 1 B, 2 = block 2 A, ...
                        block = row_idx // 2 + 1
                        row_bw = "A" if row_idx % 2 == 0 else "B"
                        if row_bw != bw:
                            continue
                        if block > num_blocks:
                            continue
                        tracks[tnum - 1].assignments.append((block, bw, code))

            return tracks

        # --- Fallback: derive from base rotation sequence in column C ---
        base_seq: dict[int, dict[str, str]] = {}  # {position: {"A": code, "B": code}}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=7, max_row=60, max_col=3, values_only=True)
        ):
            code = _str(row[2])
            if not code:
                continue

            # Try reading pos/biweek from cells first
            pos = row[0]
            biweek = _str(row[1])

            # Fall back to derivation from row index if formula cells are None
            if pos is None or biweek not in ("A", "B"):
                pos_int = row_idx // 2 + 1
                biweek = "A" if row_idx % 2 == 0 else "B"
            else:
                try:
                    pos_int = int(pos)
                except (ValueError, TypeError):
                    continue

            if pos_int not in base_seq:
                base_seq[pos_int] = {}
            base_seq[pos_int][biweek] = _normalize_rotation(code)

        if not base_seq:
            return []

        seq_len = max(base_seq.keys())
        stride = 2

        # seq_len is authoritative: it equals the number of distinct tracks
        # (header may show fewer columns than the sequence actually contains)
        actual_track_count = max(track_count, seq_len)
        tracks = [TrackTemplate(number=i + 1, label=f"Track {i + 1}")
                  for i in range(actual_track_count)]

        for tnum in range(1, actual_track_count + 1):
            for block in range(1, num_blocks + 1):
                pos = ((tnum - 1) + (block - 1) * stride) % seq_len + 1
                codes = base_seq.get(pos, {})
                for biweek in ("A", "B"):
                    code = codes.get(biweek, "")
                    if code:
                        tracks[tnum - 1].assignments.append((block, biweek, code))

        return tracks

    def read_r1_tracks(self) -> list[TrackTemplate]:
        return self._read_tracks("R1 Tracks")

    def read_r2_tracks(self) -> list[TrackTemplate]:
        return self._read_tracks("R2 Tracks")

    # ── R3-4 Recs ────────────────────────────────────────────

    def read_r34_recs_static(self, residents: list[Resident]) -> None:
        """Read only static (non-formula) data from R3-4 Recs tab.

        Reads pathway flags (cols C-F) and FSE/Top Sections fallback (col G).
        Does NOT read formula-driven columns (H: deficient sections, I-U/AB: recommended blocks).
        Use compute_r34_recs() after this to populate those fields in Python.
        """
        from schedule_maker.models.resident import FSEPrefs

        ws = self._wb["R3-4 Recs"]
        name_to_resident = {r.name: r for r in residents}

        for row in ws.iter_rows(min_row=3, values_only=True):
            name = _str(row[1]) if len(row) > 1 else ""
            if name not in name_to_resident:
                continue

            res = name_to_resident[name]

            # Parse pathway flags from columns C-F
            res.pathway = _parse_pathway(
                _str(row[2]) if len(row) > 2 else "",  # C = ESNR
                _str(row[3]) if len(row) > 3 else "",  # D = ESIR
                _str(row[4]) if len(row) > 4 else "",  # E = T32
                _str(row[5]) if len(row) > 5 else "",  # F = NRDR
            )

            # Col G (index 6): FSE/Top Sections — fallback for fse_prefs if empty
            fse_top_str = _str(row[6]) if len(row) > 6 else ""
            if fse_top_str and not res.fse_prefs and res.r_year == 4:
                specialties = [s.strip() for s in fse_top_str.split(",") if s.strip()]
                if specialties:
                    res.fse_prefs = FSEPrefs(specialties=specialties)

    def read_r34_recs(self, residents: list[Resident]) -> None:
        """Read R3-4 Recs tab and populate residents' recommended_blocks, deficient_sections, and pathways.

        Row 2 = headers: A=R#, B=Name, C=ESNR, D=ESIR, E=T32, F=NRDR,
                         G=FSE/Top Sections, H=Deficient Sections,
                         I=Vnuc, J=Smr, K=Ser, L=Sbi, M=Mnuc, N=Pcbi,
                         O=Mch, P=Mai, Q=Mus, R=Mb, S=Mucic, T=Peds,
                         U=Zir, V=Mir, W=Total, X=? ... AB=Mx, ... AD=Res/CEP
        """
        from schedule_maker.models.resident import FSEPrefs

        ws = self._wb["R3-4 Recs"]
        name_to_resident = {r.name: r for r in residents}

        # Column mapping (0-indexed from row 2 headers)
        rec_cols = {
            9: "Smr", 10: "Ser", 11: "Sbi", 12: "Mnuc",
            13: "Pcbi", 14: "Mch", 15: "Mai", 16: "Mus", 17: "Mb",
            18: "Mucic", 19: "Peds", 20: "Zir", 21: "Mir",
        }
        # Extended columns: 27=Mx blocks, 29=Res/CEP months
        MX_COL = 27
        RES_CEP_COL = 29

        for row in ws.iter_rows(min_row=3, values_only=True):
            name = _str(row[1]) if len(row) > 1 else ""
            if name not in name_to_resident:
                continue

            res = name_to_resident[name]

            # Parse pathway flags from columns C-F
            res.pathway = _parse_pathway(
                _str(row[2]) if len(row) > 2 else "",  # C = ESNR
                _str(row[3]) if len(row) > 3 else "",  # D = ESIR
                _str(row[4]) if len(row) > 4 else "",  # E = T32
                _str(row[5]) if len(row) > 5 else "",  # F = NRDR
            )

            # Col G (index 6): FSE/Top Sections — fallback for fse_prefs if empty
            fse_top_str = _str(row[6]) if len(row) > 6 else ""
            if fse_top_str and not res.fse_prefs and res.r_year == 4:
                specialties = [s.strip() for s in fse_top_str.split(",") if s.strip()]
                if specialties:
                    res.fse_prefs = FSEPrefs(specialties=specialties)

            res.deficient_sections = [
                s.strip() for s in _str(row[7]).split(",") if s.strip()
            ] if len(row) > 7 and row[7] else []

            for col_idx, rotation in rec_cols.items():
                if len(row) > col_idx and row[col_idx]:
                    val = _float(row[col_idx])
                    if val > 0:
                        res.recommended_blocks[rotation] = val

            # Mx blocks from R3-4 Recs
            if len(row) > MX_COL and row[MX_COL]:
                mx_val = _float(row[MX_COL])
                if mx_val > 0:
                    res.recommended_blocks["Mx"] = mx_val

            # Res/CEP from R3-4 Recs (fallback if not set from prefs)
            if len(row) > RES_CEP_COL and row[RES_CEP_COL]:
                res_cep_val = int(_float(row[RES_CEP_COL]))
                if res_cep_val > 0 and res.research_months == 0 and res.cep_months == 0:
                    res.research_months = res_cep_val

    # ── Preferences tab ──────────────────────────────────────

    def read_preferences_tab(self, residents: list[Resident]) -> None:
        """Read the Preferences tab from the .xlsm (manually entered data).

        Headers at row 2: R#, Name, ESNR, ESIR, T32, NRDR, then section exposures,
        then per-class preference columns, then No Call/Vac/Acad/Leave (col AA = index 26).

        0-indexed columns from values_only tuple:
          0: R#, 1: Name, 2: ESNR, 3: ESIR, 4: T32, 5: NRDR,
          6-16: historical section weeks (skip),
          17: Section Request, 18: R1 Msamp Rank, 19: Ineligible Tracks (skip),
          20: R2 Track Rank, 21: R3 Top Sections, 22: R3 Bottom Sections,
          23: R3 Zir block, 24: R3 AIRP block, 25: R4 FSE, 26: No Call dates.
        """
        from schedule_maker.models.resident import (
            AIRPPrefs, FSEPrefs, NoCallDates, Pathway,
            SamplerPrefs, SectionPrefs, TrackPrefs, ZirPrefs,
        )

        ws = self._wb["Preferences"]
        name_to_resident = {r.name: r for r in residents}
        # Also build case-insensitive lookup
        for r in residents:
            name_to_resident.setdefault(r.name.lower(), r)

        _SECTION_TO_CODES = SECTION_TO_ROTATION_CODES

        for row in ws.iter_rows(min_row=3, values_only=True):
            name = _str(row[1]) if len(row) > 1 else ""
            res = name_to_resident.get(name) or name_to_resident.get(name.lower())
            if res is None:
                continue

            r_year = int(row[0]) if len(row) > 0 and row[0] else res.r_year

            # ── Pathways (cols 2-5) — authoritative, use = not |= ──
            pathway = Pathway.NONE
            if _str(row[2]).upper() in ("X", "YES", "TRUE", "1") if len(row) > 2 and row[2] else False:
                pathway |= Pathway.ESNR
            if _str(row[3]).upper() in ("X", "YES", "TRUE", "1") if len(row) > 3 and row[3] else False:
                pathway |= Pathway.ESIR
            if _str(row[4]).upper() in ("X", "YES", "TRUE", "1") if len(row) > 4 and row[4] else False:
                pathway |= Pathway.T32
            if _str(row[5]).upper() in ("X", "YES", "TRUE", "1") if len(row) > 5 and row[5] else False:
                pathway |= Pathway.NRDR
            res.pathway = pathway

            # ── Col 17: Section Request (R4) ──
            sec_req_str = _str(row[17]) if len(row) > 17 else ""
            if sec_req_str and r_year == 4:
                scores: dict[str, int] = {}
                for abbrev in (p.strip() for p in sec_req_str.split(",") if p.strip()):
                    for code in _SECTION_TO_CODES.get(abbrev, []):
                        scores[code] = scores.get(code, 0) + 1
                if scores:
                    res.section_prefs = SectionPrefs(scores=scores)

            # ── Col 18: R1 Msamp Rank ──
            msamp_str = _str(row[18]) if len(row) > 18 else ""
            if msamp_str and r_year == 1:
                rankings: dict[str, int] = {}
                for rank_pos, code in enumerate(
                    (p.strip() for p in msamp_str.split(",") if p.strip()), 1
                ):
                    rankings[code] = rank_pos
                res.sampler_prefs = SamplerPrefs(rankings=rankings)

            # ── Col 20: R2 Track Rank ──
            track_str = _str(row[20]) if len(row) > 20 else ""
            if track_str and r_year == 2:
                rankings_t: dict[int, int] = {}
                for rank_pos, num_str in enumerate(
                    (p.strip() for p in track_str.split(",") if p.strip()), 1
                ):
                    try:
                        rankings_t[int(num_str)] = rank_pos
                    except ValueError:
                        pass
                res.track_prefs = TrackPrefs(rankings=rankings_t)

            # ── Col 21-22: R3 Top/Bottom Sections ──
            top_str = _str(row[21]) if len(row) > 21 else ""
            bottom_str = _str(row[22]) if len(row) > 22 else ""
            if (top_str or bottom_str) and r_year == 3:
                top_list = [p.strip() for p in top_str.split(",") if p.strip()]
                bottom_list = [p.strip() for p in bottom_str.split(",") if p.strip()]
                sec_scores: dict[str, int] = {}
                for i, code in enumerate(top_list):
                    sec_scores[code] = 3 - i  # +3, +2, +1
                for i, code in enumerate(bottom_list):
                    sec_scores[code] = -(3 - i)  # -3, -2, -1
                res.section_prefs = SectionPrefs(
                    top=top_list, bottom=bottom_list, scores=sec_scores,
                )

            # ── Col 23: R3 Zir block ──
            zir_str = _str(row[23]) if len(row) > 23 else ""
            if zir_str and r_year == 3:
                zir_blocks = []
                for part in zir_str.split(","):
                    try:
                        zir_blocks.append(int(part.strip()))
                    except ValueError:
                        pass
                res.zir_prefs = ZirPrefs(preferred_blocks=zir_blocks)

            # ── Col 24: R3 AIRP block ──
            airp_str = _str(row[24]) if len(row) > 24 else ""
            if airp_str and r_year == 3:
                airp_rankings: dict[str, int] = {}
                for rank_pos, session_id in enumerate(
                    (p.strip() for p in airp_str.split(",") if p.strip()), 1
                ):
                    airp_rankings[session_id] = rank_pos
                res.airp_prefs = AIRPPrefs(rankings=airp_rankings)

            # ── Col 25: R4 FSE ──
            fse_str = _str(row[25]) if len(row) > 25 else ""
            if fse_str and r_year == 4:
                if "do not want" not in fse_str.lower():
                    res.fse_prefs = FSEPrefs(
                        specialties=[s.strip() for s in fse_str.split(",") if s.strip()],
                    )

            # ── Col 26: No Call dates ──
            nocall_str = _str(row[26]) if len(row) > 26 else ""
            if nocall_str:
                raw_dates = [p.strip() for p in nocall_str.split(",") if p.strip()]
                res.no_call = NoCallDates(raw_dates=raw_dates)

    # ── Base Schedule structure ───────────────────────────────

    def read_base_schedule_structure(self) -> dict:
        """Read the Base Schedule tab structure (dates, resident rows, staffing rows).

        Returns metadata about the grid layout.
        """
        ws = self._wb["Base Schedule"]

        # Row 1-5: headers with block IDs, biweek labels, dates
        # Read date row (row 4 or 5) to get weekly dates
        dates = {}
        for row in ws.iter_rows(min_row=4, max_row=5, values_only=False):
            for cell in row:
                if hasattr(cell, "column") and cell.value and hasattr(cell.value, "year"):
                    dates[cell.column] = cell.value

        # Read resident name/PGY from rows 6+
        resident_rows = {}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=6, max_row=100, max_col=3, values_only=True),
            start=6,
        ):
            pgy = row[0] if row else None
            name = _str(row[1]) if len(row) > 1 and row[1] else ""
            if name:
                resident_rows[name] = row_idx

        # Read staffing constraint rows (101-151)
        staffing_rows = {}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=101, max_row=151, max_col=3, values_only=True),
            start=101,
        ):
            label = _str(row[0]) if row else ""
            if not label and len(row) > 1:
                label = _str(row[1])
            if label:
                staffing_rows[label] = row_idx

        return {
            "dates": dates,
            "resident_rows": resident_rows,
            "staffing_rows": staffing_rows,
        }

    def read_base_schedule_staffing(self) -> list[dict]:
        """Read staffing constraint rows from Base Schedule (rows 101-151)."""
        ws = self._wb["Base Schedule"]
        constraints = []

        for row in ws.iter_rows(min_row=101, max_row=151, values_only=True):
            label = _str(row[0]) or _str(row[1]) if len(row) > 1 else ""
            if not label:
                continue
            # Weekly staffing counts (columns D onwards = index 3+)
            weekly = [_float(row[i]) if len(row) > i else 0 for i in range(3, min(len(row), 55))]
            constraints.append({"label": label, "weekly": weekly})

        return constraints

    def read_staffing_constraints(self) -> list[StaffingConstraint]:
        """Read staffing constraints from Base Schedule rows 101-151.

        Parses: col A = min count, col B = label, col C = R# criteria.
        Labels can be either group names (e.g. "Moffitt AI") mapped via
        ROTATION_MINIMUMS, or direct rotation codes (e.g. "Mnuc", "Mb").
        """
        from schedule_maker.validation.staffing import ROTATION_MINIMUMS

        ws = self._wb["Base Schedule"]
        constraints: list[StaffingConstraint] = []

        # Build label → rotation codes from existing ROTATION_MINIMUMS
        label_codes_map: dict[str, set[str]] = {
            label: codes for label, (codes, _) in ROTATION_MINIMUMS.items()
        }
        # Additional known group labels not in ROTATION_MINIMUMS
        label_codes_map.setdefault("UC IR", {"Mir"})

        # Retired rotations — skip constraints that only reference these
        _RETIRED_ROTATIONS = {"Vch", "Vn", "Vnuc"}

        # Skip the header row (row 101: "Min", "Shift", "R# criteria")
        first_row = True
        for row in ws.iter_rows(min_row=101, max_row=151, values_only=True):
            if not row or len(row) < 2:
                continue
            if first_row:
                first_row = False
                # Check if this is a header row
                if _str(row[0]).lower() in ("min", "minimum", ""):
                    continue

            min_val = row[0]
            label = _str(row[1]) if len(row) > 1 else ""
            r_criteria_str = _str(row[2]) if len(row) > 2 else ""

            if not label:
                continue
            try:
                min_count = int(min_val) if min_val is not None else 0
            except (ValueError, TypeError):
                min_count = 0
            if min_count <= 0:
                continue

            # Parse R-year criteria: "1,2,3,4" → {1,2,3,4}
            r_years: set[int] = set()
            if r_criteria_str:
                for part in str(r_criteria_str).replace(" ", "").split(","):
                    try:
                        r_years.add(int(part))
                    except ValueError:
                        pass

            # Look up rotation codes by label
            codes: set[str] = set()

            # 1. Exact match in ROTATION_MINIMUMS group names
            if label in label_codes_map:
                codes = set(label_codes_map[label])
            else:
                # 2. Fuzzy match on group names (e.g. "ZSFG" → "ZSFG Total")
                label_lower = label.lower()
                for existing_label, existing_codes in label_codes_map.items():
                    if (existing_label.lower().startswith(label_lower)
                            or label_lower.startswith(existing_label.lower())):
                        codes = set(existing_codes)
                        break

                if not codes:
                    # 3. Label IS a rotation code (e.g. "Mnuc", "Mb", "Peds")
                    normalized = _normalize_rotation(label)
                    codes = {normalized}

            # Skip constraints that only reference retired rotations
            if codes and codes.issubset(_RETIRED_ROTATIONS):
                continue

            constraints.append(StaffingConstraint(
                label=label,
                min_count=min_count,
                rotation_codes=codes,
                r_years=r_years,
            ))

        # Post-processing: drop individual (single-code) constraints whose
        # code already appears in a multi-code group constraint.  The group
        # constraint already covers that rotation's staffing.
        group_codes: set[str] = set()
        for sc in constraints:
            if len(sc.rotation_codes) > 1:
                group_codes |= sc.rotation_codes
        constraints = [
            sc for sc in constraints
            if len(sc.rotation_codes) > 1
            or not sc.rotation_codes.issubset(group_codes)
        ]

        return constraints

    # ── NF recs ──────────────────────────────────────────────

    def read_nf_recs(self) -> dict:
        """Read NF recommendation census from NF recs tab."""
        ws = self._wb["NF recs"]
        recs = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            pgy = _str(row[0]) if row[0] else ""
            if not pgy:
                continue
            recs[pgy] = {
                "values": [_float(row[i]) if len(row) > i else 0 for i in range(1, 8)]
            }
        return recs

    # ── Night Float tab structure ────────────────────────────

    def read_night_float_structure(self) -> dict:
        """Read Night Float tab structure (similar to Base Schedule)."""
        ws = self._wb["Night Float"]

        resident_rows = {}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=6, max_row=100, max_col=3, values_only=True),
            start=6,
        ):
            name = _str(row[1]) if len(row) > 1 and row[1] else ""
            if name:
                resident_rows[name] = row_idx

        return {"resident_rows": resident_rows}

    # ── Read completed schedule assignments ───────────────────

    def read_schedule_assignments(
        self, residents: list[Resident], num_weeks: int = 52,
    ) -> dict[str, dict[int, str]]:
        """Read Base Schedule + Night Float assignments from a completed file.

        Populates res.schedule from the Base Schedule tab and returns NF
        assignments dict suitable for ScheduleGrid.nf_assignments.
        """
        name_to_resident = {r.name: r for r in residents}

        # ── Base Schedule (col D = column 4, week N at col 4 + N - 1) ──
        bs_structure = self.read_base_schedule_structure()
        bs_rows = bs_structure["resident_rows"]
        ws_bs = self._wb["Base Schedule"]
        first_data_col = 4

        for name, row_idx in bs_rows.items():
            res = name_to_resident.get(name)
            if res is None:
                continue
            row_data = list(ws_bs.iter_rows(
                min_row=row_idx, max_row=row_idx,
                min_col=first_data_col, max_col=first_data_col + num_weeks - 1,
                values_only=True,
            ))
            if not row_data:
                continue
            cells = row_data[0]
            for week in range(1, num_weeks + 1):
                idx = week - 1
                if idx < len(cells):
                    code = _normalize_rotation(_str(cells[idx]))
                    if code:
                        res.schedule[week] = code

        # ── Night Float (col F = column 6, week N at col 6 + N - 1) ──
        nf_structure = self.read_night_float_structure()
        nf_rows = nf_structure["resident_rows"]
        ws_nf = self._wb["Night Float"]
        nf_first_col = 6
        nf_codes = {"Mnf", "Snf2", "Snf", "Sx"}

        nf_assignments: dict[str, dict[int, str]] = {}
        for name, row_idx in nf_rows.items():
            if name not in name_to_resident:
                continue
            row_data = list(ws_nf.iter_rows(
                min_row=row_idx, max_row=row_idx,
                min_col=nf_first_col, max_col=nf_first_col + num_weeks - 1,
                values_only=True,
            ))
            if not row_data:
                continue
            cells = row_data[0]
            resident_nf: dict[int, str] = {}
            for week in range(1, num_weeks + 1):
                idx = week - 1
                if idx < len(cells):
                    code = _normalize_rotation(_str(cells[idx]))
                    if code and code in nf_codes:
                        resident_nf[week] = code
            if resident_nf:
                nf_assignments[name] = resident_nf

        return nf_assignments
