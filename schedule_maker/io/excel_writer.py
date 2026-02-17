"""Write schedule data back to the .xlsm workbook, preserving VBA/formulas."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl


class ExcelWriter:
    """Writes schedule assignments to the Schedule Creation .xlsm file.

    CRITICAL: Always operates on a backup copy. Never modifies the original file
    unless explicitly told to.
    """

    def __init__(self, source_path: str | Path, output_path: str | Path | None = None):
        self.source_path = Path(source_path)
        if output_path is None:
            stem = self.source_path.stem
            output_path = self.source_path.parent / f"{stem}_output.xlsm"
        self.output_path = Path(output_path)

        # Copy source to output
        shutil.copy2(self.source_path, self.output_path)

        # Open with keep_vba=True to preserve macros
        self._wb = openpyxl.load_workbook(
            str(self.output_path), keep_vba=True
        )

    def close(self):
        self._wb.close()

    def save(self):
        self._wb.save(str(self.output_path))

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.save()
        self.close()

    # ── Overview ──────────────────────────────────────────────

    def set_academic_year(self, year: int) -> None:
        """Set the target academic year in Overview tab.

        Per plan: cell B6 needs to be set to 2026 for AY 2026-2027.
        """
        ws = self._wb["Overview"]
        # Try B5 first, then B6 (varies by version)
        if ws["B5"].value and isinstance(ws["B5"].value, (int, float)):
            ws["B5"] = year
        else:
            ws["B6"] = year

    # ── Base Schedule ─────────────────────────────────────────

    def write_base_schedule(
        self,
        assignments: dict[str, dict[int, str]],
        resident_row_map: dict[str, int],
        first_data_col: int = 4,  # Column D = 4
    ) -> None:
        """Write rotation assignments to Base Schedule tab.

        Args:
            assignments: {resident_name: {week_number: rotation_code}}
            resident_row_map: {resident_name: row_number} from read_base_schedule_structure
            first_data_col: first column of weekly data (D=4)
        """
        ws = self._wb["Base Schedule"]

        for name, weeks in assignments.items():
            row = resident_row_map.get(name)
            if row is None:
                print(f"WARNING: No row found for resident {name}")
                continue

            for week_num, code in weeks.items():
                col = first_data_col + week_num - 1
                ws.cell(row=row, column=col, value=code)

    def clear_base_schedule(
        self,
        resident_row_map: dict[str, int],
        first_data_col: int = 4,
        num_weeks: int = 52,
    ) -> None:
        """Clear all base schedule assignments (right-click > Clear Contents)."""
        ws = self._wb["Base Schedule"]
        for name, row in resident_row_map.items():
            for col in range(first_data_col, first_data_col + num_weeks):
                ws.cell(row=row, column=col, value=None)

    # ── Night Float ───────────────────────────────────────────

    def write_night_float(
        self,
        nf_assignments: dict[str, dict[int, str]],
        resident_row_map: dict[str, int],
        first_data_col: int = 6,  # Column F = 6 (NF tab has extra columns)
    ) -> None:
        """Write NF overlay to Night Float tab.

        Only writes non-empty assignments. Replaces formula cells with literal strings.
        """
        ws = self._wb["Night Float"]

        for name, weeks in nf_assignments.items():
            row = resident_row_map.get(name)
            if row is None:
                print(f"WARNING: No NF row found for resident {name}")
                continue

            for week_num, code in weeks.items():
                col = first_data_col + week_num - 1
                ws.cell(row=row, column=col, value=code)

    # ── Single-cell write ─────────────────────────────────────

    def write_cell(self, sheet_name: str, row: int, col: int, value) -> None:
        """Write a single cell value to any sheet."""
        ws = self._wb[sheet_name]
        ws.cell(row=row, column=col, value=value)
