"""Parse Google Forms preference responses into Resident models."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

from schedule_maker.models.resident import (
    AIRPPrefs,
    BlockPrefs,
    FSEPrefs,
    NoCallDates,
    Pathway,
    Resident,
    SamplerPrefs,
    SectionPrefs,
    TrackPrefs,
    ZirPrefs,
)


def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _int(val) -> int:
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        # Try to extract number from "#3" format
        s = _str(val)
        if s.startswith("#"):
            try:
                return int(s[1:])
            except ValueError:
                pass
        return 0


def _parse_rank(val) -> int | None:
    """Parse a rank value like '#1', '1', or 'Top 1' into an integer."""
    s = _str(val)
    if not s:
        return None
    s = s.replace("#", "").strip()
    try:
        return int(s)
    except ValueError:
        return None


# ── Combined form column constants ─────────────────────────
_PGY_TO_RYEAR = {
    "PGY-1": 1,
    "PGY-2 (R1)": 2,
    "PGY-3 (R2)": 3,
    "PGY-4 (R3)": 4,
}

_SAMPLER_COLS: dict[int, str] = {
    4: "Nir", 5: "Mir", 6: "Msk", 7: "Mnuc", 8: "Mucic",
}

_SECTION_COLS: dict[int, str] = {
    28: "Mnuc", 29: "Mucic", 30: "Mai", 31: "Mus", 32: "Peds",
    33: "Mch", 34: "Mb", 35: "Sbi", 36: "Smr", 37: "Ser",
    38: "Vnuc", 39: "Pcbi", 40: "Zir",
}

# AIRP cols → session IDs matching AIRP_SESSIONS keys in r3_builder.py
_AIRP_COLS: dict[int, str] = {
    52: "2", 53: "3", 54: "5", 55: "9", 56: "10",
}

# Zir block preference columns (scattered across form due to Google Forms quirks).
# Cols 41-51 and 94-99 may contain Zir block headers; block numbers are parsed
# dynamically from headers.  Cols 44 (truncated "Block "), 98 ("No preference"),
# and 99 ("Row 16") are excluded by the regex.
_ZIR_COL_RANGE = list(range(41, 52)) + list(range(94, 100))


class PrefsParser:
    """Parse the Google Forms responses .xlsx into resident preferences.

    Works with the pre-cleaned per-class sheets (R1 Rotations, R2 Rotations,
    R3 Rotations, R4 Rotations, No Call Pref, CallGad Responses).
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self._wb = openpyxl.load_workbook(str(self.path), read_only=True, data_only=True)

    def close(self):
        self._wb.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    def _read_sheet_as_dicts(self, sheet_name: str) -> list[dict]:
        """Read a sheet into a list of dicts using row 1 as headers."""
        ws = self._wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [_str(h) for h in rows[0]]
        result = []
        for row in rows[1:]:
            if not any(row):
                continue
            d = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    d[header] = row[i]
            result.append(d)
        return result

    # ── R1 Preferences ────────────────────────────────────────

    def parse_r1_prefs(self, residents: list[Resident]) -> None:
        """Parse R1 Rotations sheet and update R1 residents in-place."""
        if "R1 Rotations" not in self._wb.sheetnames:
            return

        rows = self._read_sheet_as_dicts("R1 Rotations")
        name_map = {f"{r.first_name} {r.last_name}": r for r in residents if r.r_year == 1}
        # Also try "Last, First" format
        name_map.update({r.name: r for r in residents if r.r_year == 1})

        for row in rows:
            full_name = _str(row.get("Name", ""))
            if not full_name:
                first = _str(row.get("First Name", ""))
                last = _str(row.get("Last Name", ""))
                full_name = f"{first} {last}"

            res = name_map.get(full_name)
            if not res:
                # Try reverse lookup
                for key, r in name_map.items():
                    if (r.first_name in full_name and r.last_name in full_name):
                        res = r
                        break
            if not res:
                continue

            # Sampler preferences: Nir, Mir, Msk, Mnuc, Mucic columns
            rankings = {}
            for code_col in ["Nir", "Mir", "Msk", "Mnuc", "Mucic"]:
                rank = _parse_rank(row.get(code_col))
                if rank is not None:
                    rankings[code_col] = rank
            res.sampler_prefs = SamplerPrefs(rankings=rankings)

            # Vacation/Academic/Leave
            res.vacation_dates = [_str(row.get("Vac", ""))]
            res.academic_dates = [_str(row.get("Acad", ""))]
            res.leave_info = _str(row.get("Leave", ""))

    # ── R2 Preferences ────────────────────────────────────────

    def parse_r2_prefs(self, residents: list[Resident]) -> None:
        """Parse R2 Rotations sheet and update R2 residents in-place."""
        if "R2 Rotations" not in self._wb.sheetnames:
            return

        rows = self._read_sheet_as_dicts("R2 Rotations")
        name_map = self._build_name_map(residents, r_year=2)

        for row in rows:
            res = self._find_resident(row, name_map)
            if not res:
                continue

            # Track rankings from "Track Rank" column (comma-separated)
            track_rank_str = _str(row.get("Track Rank", ""))
            rankings = {}
            if track_rank_str:
                parts = [p.strip() for p in track_rank_str.split(",")]
                for rank_pos, track_num_str in enumerate(parts, 1):
                    try:
                        track_num = int(track_num_str)
                        rankings[track_num] = rank_pos
                    except ValueError:
                        pass
            else:
                # Fall back to individual track columns (F through T)
                for header, val in row.items():
                    if header.startswith("#") or (val and _str(val).startswith("#")):
                        pass  # Individual column format handled below

            res.track_prefs = TrackPrefs(rankings=rankings)

            # Pathway interest
            pathway_str = _str(row.get("Specialty Pathway Interest", ""))
            self._parse_pathway_string(res, pathway_str)

            # No-call dates
            res.no_call = NoCallDates(
                holidays=[_str(row.get("NO CALL Holiday Request", ""))],
            )
            res.vacation_dates = [_str(row.get("Vac", ""))]

    # ── R3 Preferences ────────────────────────────────────────

    def parse_r3_prefs(self, residents: list[Resident]) -> None:
        """Parse R3 Rotations sheet and update R3 residents in-place."""
        if "R3 Rotations" not in self._wb.sheetnames:
            return

        rows = self._read_sheet_as_dicts("R3 Rotations")
        name_map = self._build_name_map(residents, r_year=3)

        section_codes = [
            "Mnuc", "Mucic", "Mai", "Mus", "Peds", "Mch",
            "Mb", "Sbi", "Smr", "Ser", "Vnuc", "Pcbi", "Zir",
        ]

        for row in rows:
            res = self._find_resident(row, name_map)
            if not res:
                continue

            # Section preferences (TOP/BOTTOM scores)
            scores = {}
            top_sections = []
            bottom_sections = []
            for code in section_codes:
                val = _str(row.get(code, ""))
                if val:
                    # Values might be "TOP 1", "TOP 2", "BOTTOM 1", etc.
                    if "top" in val.lower() or val.startswith("#"):
                        rank = _parse_rank(val)
                        if rank:
                            scores[code] = 4 - rank  # TOP 1 = +3, TOP 2 = +2, TOP 3 = +1
                            top_sections.append(code)
                    elif "bottom" in val.lower():
                        rank = _parse_rank(val)
                        if rank:
                            scores[code] = -(4 - rank)  # BOTTOM 1 = -3
                            bottom_sections.append(code)

            top_str = _str(row.get("TOP Sections", ""))
            bottom_str = _str(row.get("BOTTOM Sections", ""))

            res.section_prefs = SectionPrefs(
                top=top_sections or [s.strip() for s in top_str.split(",") if s.strip()],
                bottom=bottom_sections or [s.strip() for s in bottom_str.split(",") if s.strip()],
                scores=scores,
            )

            # Zir preferences
            zir_str = _str(row.get("Zir block pref", ""))
            zir_blocks = []
            for part in zir_str.split(","):
                try:
                    zir_blocks.append(int(part.strip()))
                except ValueError:
                    pass
            res.zir_prefs = ZirPrefs(preferred_blocks=zir_blocks)

            # AIRP preferences
            airp_rank_str = _str(row.get("AIRP block rank", ""))
            airp_rankings = {}
            if airp_rank_str:
                parts = [p.strip() for p in airp_rank_str.split(",")]
                for rank_pos, session_id in enumerate(parts, 1):
                    try:
                        # Session IDs might be like "2", "3+4", "4+5", "9", "10"
                        airp_rankings[session_id] = rank_pos
                    except (ValueError, TypeError):
                        pass

            airp_group = _str(row.get("AIRP group", ""))
            res.airp_prefs = AIRPPrefs(
                rankings=airp_rankings,
                group_requests=[g.strip() for g in airp_group.split(",") if g.strip()],
            )

            # Pathway
            pathway_str = _str(row.get("Specialty Pathway Interest", ""))
            self._parse_pathway_string(res, pathway_str)

            # No-call and vacation
            res.no_call = NoCallDates(
                holidays=[_str(row.get("NO CALL Holiday Request", ""))],
            )
            res.vacation_dates = [_str(row.get("Vac", ""))]
            res.academic_dates = [_str(row.get("Acad", ""))]
            res.leave_info = _str(row.get("Leave", ""))

    # ── R4 Preferences ────────────────────────────────────────

    def parse_r4_prefs(self, residents: list[Resident]) -> None:
        """Parse R4 Rotations sheet and update R4 residents in-place."""
        if "R4 Rotations" not in self._wb.sheetnames:
            return

        rows = self._read_sheet_as_dicts("R4 Rotations")
        name_map = self._build_name_map(residents, r_year=4)

        for row in rows:
            res = self._find_resident(row, name_map)
            if not res:
                continue

            # FSE preferences
            fse_str = _str(row.get("FSE", ""))
            fse_org = _str(row.get("FSE/Rotation Pref", ""))
            dist_pref = _str(row.get("Distribution Pref", ""))
            res.fse_prefs = FSEPrefs(
                specialties=[s.strip() for s in fse_str.split(",") if s.strip()],
                organization=dist_pref or fse_org,
            )

            # Research/CEP months
            res.research_months = _int(row.get("Research Months", 0))
            res.cep_months = _int(row.get("CEP Months", 0))

            # Pathway flags (explicit columns)
            if _str(row.get("T32", "")).lower() in ("x", "yes", "true"):
                res.pathway |= Pathway.T32
            if _str(row.get("ESIR", "")).lower() in ("x", "yes", "true"):
                res.pathway |= Pathway.ESIR
            if _str(row.get("NRDR", "")).lower() in ("x", "yes", "true"):
                res.pathway |= Pathway.NRDR
            if _str(row.get("ESNR", "")).lower() in ("x", "yes", "true"):
                res.pathway |= Pathway.ESNR

            # Section preference counts (individual columns like "Mai", "Mus", etc.)
            section_counts = {}
            for code in ["Mai", "Mus", "Mb", "Ser", "Mch", "Mucic", "Peds", "Smr"]:
                val = row.get(code)
                if val is not None:
                    section_counts[code] = _int(val)
            if section_counts:
                res.section_prefs = SectionPrefs(scores=section_counts)

            # Block preferences (columns with numbers 1-13 as headers)
            block_prefs = {}
            for block_num in range(1, 14):
                val = _str(row.get(str(block_num), ""))
                if val:
                    block_prefs[block_num] = val
            if block_prefs:
                res.block_prefs = BlockPrefs(assignments=block_prefs)

            # No-call and vacation
            res.no_call = NoCallDates(
                holidays=[_str(row.get("NO CALL Holiday Request", ""))],
            )
            res.vacation_dates = [_str(row.get("Vac", ""))]
            res.academic_dates = [_str(row.get("Acad", ""))]
            res.leave_info = _str(row.get("Leave", ""))

    # ── No Call Pref (consolidated) ───────────────────────────

    def parse_no_call_prefs(self, residents: list[Resident]) -> None:
        """Parse No Call Pref sheet for consolidated no-call/NF assignment data."""
        if "No Call Pref" not in self._wb.sheetnames:
            return

        rows = self._read_sheet_as_dicts("No Call Pref")
        name_map = {f"{r.first_name} {r.last_name}": r for r in residents}
        name_map.update({r.name: r for r in residents})

        for row in rows:
            full_name = _str(row.get("Name", ""))
            res = name_map.get(full_name)
            if not res:
                for key, r in name_map.items():
                    if r.last_name and r.last_name in full_name:
                        res = r
                        break
            if not res:
                continue

            # "NO NF ASSIGNMENTS" is the comma-separated formatted dates
            no_nf_str = _str(row.get("NO NF ASSIGNMENTS", ""))
            if no_nf_str:
                # Format: "Name:date1,date2,date3"
                if ":" in no_nf_str:
                    no_nf_str = no_nf_str.split(":", 1)[1]
                # Parse MM/DD dates - store as raw strings for now
                raw_parts = [p.strip() for p in no_nf_str.split(",") if p.strip()]
                res.no_call.raw_dates = []
                for part in raw_parts:
                    # These are in M/D format
                    res.no_call.raw_dates.append(part)

    # ── Parse all ─────────────────────────────────────────────

    def parse_all(self, residents: list[Resident]) -> None:
        """Parse all preference sheets and update residents in-place.

        Auto-detects format: if a combined "Form Responses 1" sheet exists
        and no per-class sheets are present, uses the combined parser.
        Otherwise falls back to the per-class sheet parser.
        """
        has_combined = "Form Responses 1" in self._wb.sheetnames
        has_per_class = any(
            s in self._wb.sheetnames
            for s in ("R1 Rotations", "R2 Rotations", "R3 Rotations", "R4 Rotations")
        )

        if has_combined and not has_per_class:
            self._parse_combined_form(residents)
        else:
            self.parse_r1_prefs(residents)
            self.parse_r2_prefs(residents)
            self.parse_r3_prefs(residents)
            self.parse_r4_prefs(residents)
            self.parse_no_call_prefs(residents)

    # ── Combined form parsing ─────────────────────────────────

    def _parse_combined_form(self, residents: list[Resident]) -> None:
        """Parse the unified 'Form Responses 1' sheet (2026-2027+ format)."""
        ws = self._wb["Form Responses 1"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return

        headers = rows[0]

        # Build Zir col → block number mapping from headers
        zir_col_map: dict[int, int] = {}
        for i in _ZIR_COL_RANGE:
            if i >= len(headers) or headers[i] is None:
                continue
            m = re.search(r"Block\s+(\d+)", str(headers[i]))
            if m:
                zir_col_map[i] = int(m.group(1))

        # Build case-insensitive name lookup for all residents
        name_map: dict[str, Resident] = {}
        for r in residents:
            name_map[f"{r.first_name} {r.last_name}".lower()] = r
            name_map[r.name.lower()] = r
            name_map[f"{r.last_name}, {r.first_name}".lower()] = r

        # Deduplicate: keep only the latest response per (first, last)
        latest: dict[tuple[str, str], tuple] = {}
        for row in rows[1:]:
            if not any(row):
                continue
            first = _str(row[1]).strip()
            last = _str(row[2]).strip()
            if not first or not last:
                continue
            pgy_str = _str(row[3])
            if pgy_str not in _PGY_TO_RYEAR:
                continue
            key = (first.lower(), last.lower())
            ts = row[0]
            if key not in latest or (
                ts and (latest[key][0] is None or ts > latest[key][0])
            ):
                latest[key] = (ts, row)

        # Process each deduplicated response
        for (_first_l, _last_l), (_ts, row) in latest.items():
            first = _str(row[1]).strip()
            last = _str(row[2]).strip()
            r_year = _PGY_TO_RYEAR[_str(row[3])]

            res = self._find_resident_combined(first, last, name_map)
            if not res:
                continue

            if r_year == 1:
                self._parse_combined_r1(row, res)
            elif r_year == 2:
                self._parse_combined_r2(row, res)
            elif r_year == 3:
                self._parse_combined_r3(row, res, zir_col_map)
            elif r_year == 4:
                self._parse_combined_r4(row, res)

            self._parse_combined_shared(row, res)

    @staticmethod
    def _find_resident_combined(
        first: str, last: str, name_map: dict[str, Resident],
    ) -> Resident | None:
        """Match a combined-form name to a roster resident."""
        for key in (
            f"{first} {last}".lower(),
            f"{last}, {first}".lower(),
        ):
            if key in name_map:
                return name_map[key]
        # Fall back to case-insensitive last-name match
        last_lower = last.lower()
        for res in name_map.values():
            if res.last_name.lower() == last_lower:
                return res
        return None

    def _parse_combined_r1(self, row: tuple, res: Resident) -> None:
        """Extract R1 sampler rankings and pathway interest."""
        rankings = {}
        for col, code in _SAMPLER_COLS.items():
            rank = _parse_rank(row[col] if col < len(row) else None)
            if rank is not None:
                rankings[code] = rank
        if rankings:
            res.sampler_prefs = SamplerPrefs(rankings=rankings)

        pathway_str = _str(row[9] if 9 < len(row) else "")
        if pathway_str:
            self._parse_pathway_string(res, pathway_str)

    def _parse_combined_r2(self, row: tuple, res: Resident) -> None:
        """Extract R2 track rankings and pathway interest."""
        rankings = {}
        for col_idx in range(11, 26):
            track_num = col_idx - 10  # col 11 → Track 1, …, col 25 → Track 15
            rank = _parse_rank(row[col_idx] if col_idx < len(row) else None)
            if rank is not None:
                rankings[track_num] = rank
        if rankings:
            res.track_prefs = TrackPrefs(rankings=rankings)

        # R2 pathway (same question as R1, col 9)
        pathway_str = _str(row[9] if 9 < len(row) else "")
        if pathway_str:
            self._parse_pathway_string(res, pathway_str)

    def _parse_combined_r3(
        self, row: tuple, res: Resident, zir_col_map: dict[int, int],
    ) -> None:
        """Extract R3 section prefs, Zir block prefs, AIRP rankings, pathway."""
        # Section preferences (cols 28-40): #1-3 = top, #11-13 = bottom
        scores: dict[str, int] = {}
        top_sections: list[str] = []
        bottom_sections: list[str] = []
        for col, code in _SECTION_COLS.items():
            rank = _parse_rank(row[col] if col < len(row) else None)
            if rank is None:
                continue
            if 1 <= rank <= 3:
                scores[code] = 4 - rank        # #1→+3, #2→+2, #3→+1
                top_sections.append(code)
            elif 11 <= rank <= 13:
                scores[code] = -(14 - rank)     # #11→-3, #12→-2, #13→-1
                bottom_sections.append(code)

        res.section_prefs = SectionPrefs(
            top=top_sections,
            bottom=bottom_sections,
            scores=scores,
        )

        # Zir block preferences — collect ranked blocks, sort by rank
        block_ranks: dict[int, int] = {}
        for col, block_num in sorted(zir_col_map.items()):
            rank = _parse_rank(row[col] if col < len(row) else None)
            if rank is not None and block_num not in block_ranks:
                block_ranks[block_num] = rank
        ranked = sorted(block_ranks.items(), key=lambda x: x[1])
        res.zir_prefs = ZirPrefs(preferred_blocks=[b for b, _ in ranked])

        # AIRP session rankings (cols 52-56)
        airp_rankings: dict[str, int] = {}
        for col, session_id in _AIRP_COLS.items():
            rank = _parse_rank(row[col] if col < len(row) else None)
            if rank is not None:
                airp_rankings[session_id] = rank
        res.airp_prefs = AIRPPrefs(rankings=airp_rankings)

        # Pathway (col 26)
        pathway_str = _str(row[26] if 26 < len(row) else "")
        if pathway_str:
            self._parse_pathway_string(res, pathway_str)

    def _parse_combined_r4(self, row: tuple, res: Resident) -> None:
        """Extract R4 pathway, FSE, research/CEP."""
        # Pathway (cols 57-59)
        pursuing = _str(row[57] if 57 < len(row) else "")
        which = _str(row[58] if 58 < len(row) else "")
        t32 = _str(row[59] if 59 < len(row) else "")

        if pursuing.lower() == "yes" and which:
            self._parse_pathway_string(res, which)
        if t32.lower() == "yes":
            res.pathway |= Pathway.T32

        # FSE (cols 62-63)
        fse_choice = _str(row[62] if 62 < len(row) else "")
        fse_org = _str(row[63] if 63 < len(row) else "")
        if fse_choice:
            res.fse_prefs = FSEPrefs(
                specialties=[s.strip() for s in fse_choice.split(",") if s.strip()],
                organization=fse_org,
            )

        # Research / CEP (cols 64-65)
        for col, attr in ((64, "research_months"), (65, "cep_months")):
            val = row[col] if col < len(row) else None
            if val is not None:
                try:
                    setattr(res, attr, int(float(val)))
                except (ValueError, TypeError):
                    pass

    def _parse_combined_shared(self, row: tuple, res: Resident) -> None:
        """Extract no-call weekends, vacation, academic, leave."""
        # No-call weekends (cols 70-71) — datetime objects from Google Forms
        raw_dates: list[str] = []
        for col in (70, 71):
            val = row[col] if col < len(row) else None
            if val is None:
                continue
            if isinstance(val, datetime):
                raw_dates.append(f"{val.month}/{val.day}")
            else:
                s = _str(val)
                if s:
                    raw_dates.append(s)

        # Holiday preferences (cols 72-73)
        holiday_free = _str(row[73] if 73 < len(row) else "")
        holidays = [holiday_free] if holiday_free else []

        res.no_call = NoCallDates(raw_dates=raw_dates, holidays=holidays)

        # Vacation, academic, leave (cols 77-79)
        vac = _str(row[77] if 77 < len(row) else "")
        acad = _str(row[78] if 78 < len(row) else "")
        leave = _str(row[79] if 79 < len(row) else "")

        if vac:
            res.vacation_dates = [vac]
        if acad:
            res.academic_dates = [acad]
        if leave:
            res.leave_info = leave

    # ── Helpers ───────────────────────────────────────────────

    def _build_name_map(self, residents: list[Resident], r_year: int) -> dict[str, Resident]:
        """Build a name lookup map for residents of a given year."""
        result = {}
        for r in residents:
            if r.r_year != r_year:
                continue
            result[f"{r.first_name} {r.last_name}"] = r
            result[r.name] = r
            result[f"{r.last_name}, {r.first_name}"] = r
        return result

    def _find_resident(self, row: dict, name_map: dict[str, Resident]) -> Resident | None:
        """Try to find a resident from a preference row using various name formats."""
        full_name = _str(row.get("Full Name", "")) or _str(row.get("Name", ""))
        if not full_name:
            first = _str(row.get("First Name", ""))
            last = _str(row.get("Last Name", ""))
            full_name = f"{first} {last}"

        res = name_map.get(full_name)
        if res:
            return res

        # Try partial matching
        for key, r in name_map.items():
            if r.last_name and r.last_name in full_name:
                return r

        return None

    @staticmethod
    def _parse_pathway_string(res: Resident, pathway_str: str) -> None:
        """Parse a pathway interest string like 'ESIR, T32' into pathway flags."""
        s = pathway_str.upper()
        if "ESIR" in s:
            res.pathway |= Pathway.ESIR
        if "ESNR" in s:
            res.pathway |= Pathway.ESNR
        if "T32" in s:
            res.pathway |= Pathway.T32
        if "NR/DR" in s or "NRDR" in s or "NR-DR" in s:
            res.pathway |= Pathway.NRDR
