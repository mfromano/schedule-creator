"""Write parsed preference data into the Preferences tab of the .xlsm file."""

from __future__ import annotations

import shutil
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from schedule_maker.models.resident import Resident


# Long-form Google Forms names → Key-sheet codes (column C)
_LONG_TO_KEY: dict[str, list[str]] = {
    "Abdominal Imaging": ["AI"],
    "Chest/Cardiac": ["Chest", "Cardiac"],
    "Neuroradiology": ["Neuro"],
    "Interventional Radiology": ["IR"],
    "Breast": ["Breast"],
    "MSK": ["MSK"],
    "Pediatric Radiology": ["Peds"],
    "Nuclear Medicine": ["NucMed"],
    "Ultrasound": ["US"],
}

# Flat version for single-value lookups (first code wins)
_LONG_TO_KEY_FLAT: dict[str, str] = {
    long: codes[0] for long, codes in _LONG_TO_KEY.items()
}


# ── Column indices (1-based) in the Preferences tab ──────
_COL_SECTION_REQUEST = 18   # R
_COL_MSAMP_RANK = 19        # S
_COL_INELIGIBLE_TRACKS = 20 # T  (left empty)
_COL_TRACK_RANK = 21        # U
_COL_TOP_SECTIONS = 22      # V
_COL_BOTTOM_SECTIONS = 23   # W
_COL_ZIR_BLOCK = 24         # X
_COL_AIRP_BLOCK = 25        # Y
_COL_FSE = 26               # Z
_COL_NO_CALL = 27           # AA
_COL_AIRP_GROUP = 28        # AB
_COL_PATHWAY_ORG = 29       # AC
_COL_HOLIDAY_HISTORY = 30   # AD
_COL_HOLIDAY_WORK_PREF = 31 # AE

# Form columns for section requests (0-based)
_SECTION_REQUEST_COLS = {2: 10, 3: 27, 4: 61}  # r_year → form column index


def write_preferences(
    source_path: Path,
    residents: list[Resident],
    output_path: Path | None = None,
    nocall_datetimes: dict[str, list[datetime]] | None = None,
    section_requests: dict[str, str] | None = None,
) -> Path:
    """Write resident preferences into the Preferences tab of the .xlsm.

    Copies source to output, writes preference columns R-AA, saves.
    Returns the output path.
    """
    source_path = Path(source_path)
    if output_path is None:
        output_path = source_path.parent / f"{source_path.stem}_with_prefs.xlsm"
    output_path = Path(output_path)

    shutil.copy2(source_path, output_path)

    wb = openpyxl.load_workbook(str(output_path), keep_vba=True)
    ws = wb["Preferences"]

    # Build name → row mapping from column B (rows 3-62)
    name_to_row: dict[str, int] = {}
    for row_idx in range(3, 63):
        cell_val = ws.cell(row=row_idx, column=2).value
        if cell_val:
            name_to_row[str(cell_val).strip()] = row_idx

    # Build resident lookup by name
    res_by_name: dict[str, Resident] = {}
    for r in residents:
        res_by_name[r.name] = r
        # Also index by "Last, First" variants
        alt = f"{r.last_name}, {r.first_name}"
        res_by_name[alt] = r

    # Write headers for new columns (rows 1-2)
    ws.cell(row=1, column=_COL_AIRP_GROUP, value="R3")
    ws.cell(row=2, column=_COL_AIRP_GROUP, value="AIRP group")
    ws.cell(row=1, column=_COL_PATHWAY_ORG, value="R4")
    ws.cell(row=2, column=_COL_PATHWAY_ORG, value="Pathway org")
    ws.cell(row=1, column=_COL_HOLIDAY_HISTORY, value="All")
    ws.cell(row=2, column=_COL_HOLIDAY_HISTORY, value="Holiday call history")
    ws.cell(row=1, column=_COL_HOLIDAY_WORK_PREF, value="All")
    ws.cell(row=2, column=_COL_HOLIDAY_WORK_PREF, value="Holiday work pref")

    nocall_datetimes = nocall_datetimes or {}
    section_requests = section_requests or {}

    matched = 0
    for xl_name, row_idx in name_to_row.items():
        res = res_by_name.get(xl_name)
        if res is None:
            # Try case-insensitive match
            xl_lower = xl_name.lower()
            for key, r in res_by_name.items():
                if key.lower() == xl_lower:
                    res = r
                    break
        if res is None:
            print(f"[prefs_writer] WARNING: no form data for '{xl_name}' (row {row_idx})")
            continue

        matched += 1

        # Column R: Section Request
        sec_req = section_requests.get(res.name, "")
        if sec_req:
            ws.cell(row=row_idx, column=_COL_SECTION_REQUEST, value=sec_req)

        # Column S: Msamp Rank (R1 only)
        if res.r_year == 1:
            val = _format_sampler_rank(res)
            if val:
                ws.cell(row=row_idx, column=_COL_MSAMP_RANK, value=val)

        # Column U: Track Rank (R2 only)
        if res.r_year == 2:
            val = _format_track_rank(res)
            if val:
                ws.cell(row=row_idx, column=_COL_TRACK_RANK, value=val)

        # Column V: Top Sections (R3 only)
        if res.r_year == 3:
            val = _format_top_sections(res)
            if val:
                ws.cell(row=row_idx, column=_COL_TOP_SECTIONS, value=val)

        # Column W: Bottom Sections (R3 only)
        if res.r_year == 3:
            val = _format_bottom_sections(res)
            if val:
                ws.cell(row=row_idx, column=_COL_BOTTOM_SECTIONS, value=val)

        # Column X: Zir block (R3 only)
        if res.r_year == 3:
            val = _format_zir_blocks(res)
            if val:
                ws.cell(row=row_idx, column=_COL_ZIR_BLOCK, value=val)

        # Column Y: AIRP block (R3 only)
        if res.r_year == 3:
            val = _format_airp_rank(res)
            if val:
                ws.cell(row=row_idx, column=_COL_AIRP_BLOCK, value=val)

        # Column Z: FSE (R4 only)
        if res.r_year == 4:
            val = _format_fse(res)
            if val:
                ws.cell(row=row_idx, column=_COL_FSE, value=val)

        # Column AA: No Call, Vac, Acad, Leave
        val = _format_no_call_dates(res, nocall_datetimes.get(res.name))
        if val:
            ws.cell(row=row_idx, column=_COL_NO_CALL, value=val)

        # Column AB: AIRP group (R3 only)
        if res.r_year == 3:
            val = _format_airp_group(res)
            if val:
                ws.cell(row=row_idx, column=_COL_AIRP_GROUP, value=val)

        # Column AC: Pathway org (R4 only)
        if res.r_year == 4:
            val = _format_pathway_org(res)
            if val:
                ws.cell(row=row_idx, column=_COL_PATHWAY_ORG, value=val)

        # Column AD: Holiday call history (all years)
        val = _format_holiday_history(res)
        if val:
            ws.cell(row=row_idx, column=_COL_HOLIDAY_HISTORY, value=val)

        # Column AE: Holiday work pref (all years)
        if res.no_call.holiday_work_pref:
            ws.cell(row=row_idx, column=_COL_HOLIDAY_WORK_PREF, value=res.no_call.holiday_work_pref)

    print(f"[prefs_writer] Wrote preferences for {matched}/{len(name_to_row)} residents")

    wb.save(str(output_path))
    wb.close()
    return output_path


# ── Serialization helpers ────────────────────────────────


def _format_sampler_rank(res: Resident) -> str:
    """SamplerPrefs rankings → sorted by rank, comma-separated rotation names."""
    if res.sampler_prefs is None or not res.sampler_prefs.rankings:
        return ""
    # rankings: {rotation_code: rank}, sort by rank ascending
    sorted_items = sorted(res.sampler_prefs.rankings.items(), key=lambda x: x[1])
    return ",".join(code for code, _rank in sorted_items)


def _format_track_rank(res: Resident) -> str:
    """TrackPrefs rankings → sorted by rank, comma-separated track numbers."""
    if res.track_prefs is None or not res.track_prefs.rankings:
        return ""
    # rankings: {track_number: rank}, sort by rank ascending
    sorted_items = sorted(res.track_prefs.rankings.items(), key=lambda x: x[1])
    return ",".join(str(track) for track, _rank in sorted_items)


def _format_top_sections(res: Resident) -> str:
    """SectionPrefs.top → comma-separated codes."""
    if res.section_prefs is None or not res.section_prefs.top:
        return ""
    return ",".join(res.section_prefs.top)


def _format_bottom_sections(res: Resident) -> str:
    """SectionPrefs.bottom → comma-separated codes."""
    if res.section_prefs is None or not res.section_prefs.bottom:
        return ""
    return ",".join(res.section_prefs.bottom)


def _format_zir_blocks(res: Resident) -> str:
    """ZirPrefs → comma-separated block numbers."""
    if res.zir_prefs is None or not res.zir_prefs.preferred_blocks:
        return ""
    return ",".join(str(b) for b in res.zir_prefs.preferred_blocks)


def _format_airp_rank(res: Resident) -> str:
    """AIRPPrefs rankings → sorted by rank, comma-separated session IDs."""
    if res.airp_prefs is None or not res.airp_prefs.rankings:
        return ""
    # rankings: {session_id: rank}, sort by rank ascending
    sorted_items = sorted(res.airp_prefs.rankings.items(), key=lambda x: x[1])
    return ",".join(session_id for session_id, _rank in sorted_items)


def _format_fse(res: Resident) -> str:
    """FSEPrefs → comma-separated Key-sheet codes."""
    if res.fse_prefs is None or not res.fse_prefs.specialties:
        return ""
    codes: list[str] = []
    for spec in res.fse_prefs.specialties:
        mapped = _LONG_TO_KEY.get(spec)
        if mapped:
            codes.extend(mapped)
        else:
            codes.append(spec)
    return ",".join(codes)


def _format_airp_group(res: Resident) -> str:
    """AIRPPrefs.group_requests → comma-separated names."""
    if res.airp_prefs is None or not res.airp_prefs.group_requests:
        return ""
    return ", ".join(res.airp_prefs.group_requests)


def _format_pathway_org(res: Resident) -> str:
    """FSEPrefs.organization → pathway organization preference."""
    if res.fse_prefs is None or not res.fse_prefs.organization:
        return ""
    return res.fse_prefs.organization


def _format_holiday_history(res: Resident) -> str:
    """NoCallDates.holiday_history → semicolon-separated history entries."""
    if not res.no_call.holiday_history:
        return ""
    non_empty = [h for h in res.no_call.holiday_history if h]
    if not non_empty:
        return ""
    return "; ".join(res.no_call.holiday_history)


def _expand_nocall_weekend(friday: datetime) -> list[str]:
    """Expand a no-call weekend Friday to Fri, Sat, Sun, Mon as M/D strings."""
    dates = []
    for offset in range(4):
        dt = friday + timedelta(days=offset)
        dates.append(f"{dt.month}/{dt.day}")
    return dates


def _format_no_call_dates(
    res: Resident, nocall_dts: list[datetime] | None
) -> str:
    """Merge expanded no-call weekends with vacation/academic dates.

    nocall_dts: raw datetime objects from form cols 70-71 (Fridays to expand).
    Also includes res.vacation_dates and res.academic_dates.
    Output: comma-separated M/D strings.
    """
    all_dates: list[str] = []

    # Expand no-call weekend datetimes (Fri→Mon)
    if nocall_dts:
        for dt in nocall_dts:
            all_dates.extend(_expand_nocall_weekend(dt))

    # Add vacation dates
    for vac in res.vacation_dates:
        if vac:
            for part in vac.split(","):
                part = part.strip()
                if part:
                    all_dates.append(part)

    # Add academic dates
    for acad in res.academic_dates:
        if acad:
            for part in acad.split(","):
                part = part.strip()
                if part:
                    all_dates.append(part)

    # Add leave info if it looks like dates
    if res.leave_info:
        for part in res.leave_info.split(","):
            part = part.strip()
            if "/" in part:
                all_dates.append(part)

    return ",".join(all_dates)
