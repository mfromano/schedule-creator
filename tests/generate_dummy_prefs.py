"""Generate a dummy preferences .xlsx file for 2026-2027 testing.

Creates the multi-sheet format expected by prefs_parser.py with
randomized but realistic preferences for all 60 active residents.
"""
from __future__ import annotations

import random
import openpyxl

random.seed(42)  # Reproducible

# Read the 2026-2027 roster from the schedule file
# Layout: A=Current PGY, B=Resident, C=ESNR, D=ESIR, E=T32, F=NRDR
src = openpyxl.load_workbook(
    "Schedule Creation (2026-2027).xlsm", data_only=True, read_only=True,
)
ws = src["Historical"]

residents = []
for row in ws.iter_rows(min_row=3, max_row=80, max_col=6, values_only=True):
    current_pgy = row[0]
    name = row[1]
    if not name or not current_pgy:
        continue
    try:
        pgy = int(current_pgy) + 1  # Increment for target year
    except (ValueError, TypeError):
        continue
    r_year = pgy - 1
    if r_year < 1 or r_year > 4:
        continue
    residents.append({
        "name": str(name).strip(),
        "r_year": r_year,
        "pgy": pgy,
        "esnr": "x" if row[2] else "",
        "esir": "x" if row[3] else "",
        "t32": "x" if row[4] else "",
        "nrdr": "x" if row[5] else "",
    })
src.close()

# Parse name into first/last
for r in residents:
    parts = r["name"].split(",", 1)
    r["last_name"] = parts[0].strip()
    r["first_name"] = parts[1].strip() if len(parts) > 1 else ""
    r["full_name"] = f"{r['first_name']} {r['last_name']}"

# Separate by year
r1s = [r for r in residents if r["r_year"] == 1]
r2s = [r for r in residents if r["r_year"] == 2]
r3s = [r for r in residents if r["r_year"] == 3]
r4s = [r for r in residents if r["r_year"] == 4]

print(f"R1: {len(r1s)}, R2: {len(r2s)}, R3: {len(r3s)}, R4: {len(r4s)}")

# Create workbook
wb = openpyxl.Workbook()

# === R1 Rotations ===
ws_r1 = wb.active
ws_r1.title = "R1 Rotations"
r1_headers = ["Date", "First Name", "Last Name", "PGY", "Nir", "Mir", "Msk", "Mnuc", "Mucic",
              "Name", "Msamp Ranking", "Tentative Msamp", "Vac", "Acad", "Leave", "Comment"]
ws_r1.append(r1_headers)

for r in r1s:
    ranks = random.sample(range(1, 6), 5)
    row = [
        "2026-02-15",
        r["first_name"],
        r["last_name"],
        r["pgy"],
    ]
    for rank in ranks:
        row.append(f"#{rank}")
    row.extend([
        r["full_name"],
        "",
        "",
        "10/15, 10/16, 10/17, 3/20, 3/21",
        "11/10, 11/11",
        "",
        "",
    ])
    ws_r1.append(row)

# === R2 Rotations ===
ws_r2 = wb.create_sheet("R2 Rotations")
r2_headers = ["First Name", "Last Name", "Full Name", "Specialty Pathway Interest", "Add'l Weak",
             "Track Rank", "NO CALL Weekend Request", "NO CALL Week Request",
             "NO CALL Holiday Request", "Call Holiday Preference", "Vac", "Acad", "Leave", "Comment"]
ws_r2.append(r2_headers)

num_tracks = len(r2s)
for r in r2s:
    track_order = random.sample(range(1, num_tracks + 1), num_tracks)
    track_rank_str = ", ".join(str(t) for t in track_order)

    pathway = ""
    if r["esir"]:
        pathway = "ESIR"
    elif r["esnr"]:
        pathway = "ESNR"

    row = [
        r["first_name"],
        r["last_name"],
        r["full_name"],
        pathway or "None",
        "",
        track_rank_str,
        "",
        "",
        random.choice(["Christmas", "Thanksgiving", "New Years", ""]),
        "",
        "12/20, 12/21, 12/22, 4/5, 4/6",
        "9/15",
        "",
        "",
    ]
    ws_r2.append(row)

# === R3 Rotations ===
ws_r3 = wb.create_sheet("R3 Rotations")
section_codes = ["Mnuc", "Mucic", "Mai", "Mus", "Peds", "Mch", "Mb", "Sbi", "Smr", "Ser", "Vnuc", "Pcbi", "Zir"]
r3_headers = (["First Name", "Last Name", "Full Name", "Specialty Pathway Interest", "Add'l Weak"]
              + section_codes
              + ["TOP Sections", "BOTTOM Sections", "Zir block pref", "AIRP block rank", "AIRP group",
                 "NO CALL Holiday Request", "NO CALL Weekend Request", "Vac", "Acad", "Leave", "Comment"])
ws_r3.append(r3_headers)

airp_sessions = ["2", "3+4", "4+5", "9", "10"]

for r in r3s:
    pathway = ""
    if r["esir"]:
        pathway = "ESIR"
    elif r["esnr"]:
        pathway = "ESNR"
    elif r["nrdr"]:
        pathway = "NR/DR"
    elif r["t32"]:
        pathway = "T32"

    row = [
        r["first_name"],
        r["last_name"],
        r["full_name"],
        pathway or "None",
        "",
    ]

    section_ranks = random.sample(range(1, 14), 13)
    for rank in section_ranks:
        row.append(f"#{rank}")

    top_indices = sorted(range(13), key=lambda i: section_ranks[i])[:3]
    bottom_indices = sorted(range(13), key=lambda i: -section_ranks[i])[:3]
    top_str = ", ".join(section_codes[i] for i in top_indices)
    bottom_str = ", ".join(section_codes[i] for i in bottom_indices)

    zir_blocks = random.sample(range(1, 7), 3)
    zir_str = ", ".join(str(b) for b in sorted(zir_blocks))

    airp_order = random.sample(airp_sessions, len(airp_sessions))
    airp_str = ", ".join(airp_order)

    other_r3s = [x for x in r3s if x["name"] != r["name"]]
    group_mate = random.choice(other_r3s)["full_name"] if random.random() < 0.3 else ""

    row.extend([
        top_str,
        bottom_str,
        zir_str,
        airp_str,
        group_mate,
        random.choice(["Christmas", "Thanksgiving", "New Years", ""]),
        "",
        "3/15, 3/16, 3/17, 6/1, 6/2",
        "10/5",
        "",
        "",
    ])
    ws_r3.append(row)

# === R4 Rotations ===
ws_r4 = wb.create_sheet("R4 Rotations")
r4_section_codes = ["Mai", "Mus", "Mb", "Ser", "Mch", "Mucic", "Peds", "Smr"]
r4_headers = (["First Name", "Last Name", "Full Name",
               "T32", "ESIR", "NRDR", "ESNR",
               "Section Pref", "FSE", "FSE/Rotation Pref", "Distribution Pref",
               "Research Months", "CEP Months"]
              + [str(b) for b in range(1, 14)]
              + r4_section_codes
              + ["NO CALL Holiday Request", "Vac", "Acad", "Leave", "Comment"])
ws_r4.append(r4_headers)

fse_options = ["Abdominal Imaging", "Breast", "Neuroradiology", "Chest", "MSK", ""]

for r in r4s:
    if r["t32"]:
        research = 0
        cep = 0
    else:
        research = random.choice([0, 0, 0, 1, 1, 2])
        cep = random.choice([0, 0, 0, 1])

    fse = random.choice(fse_options)
    if r["nrdr"] or r["esir"]:
        fse = ""

    row = [
        r["first_name"],
        r["last_name"],
        r["full_name"],
        r["t32"],
        r["esir"],
        r["nrdr"],
        r["esnr"],
        "",
        fse,
        "Contiguous" if fse else "",
        "Sequential" if fse else "",
        research,
        cep,
    ]

    for b in range(1, 14):
        row.append("")

    for code in r4_section_codes:
        row.append(random.randint(1, 5))

    row.extend([
        random.choice(["Christmas", "Thanksgiving", "New Years", ""]),
        "11/25, 11/26, 5/10, 5/11",
        "2/15",
        "",
        "",
    ])
    ws_r4.append(row)

# === No Call Pref ===
ws_nc = wb.create_sheet("No Call Pref")
nc_headers = ["First Name", "Last Name", "Current PGY", "Name", "R#",
              "Prior Holiday Call 2023-2024", "Prior Holiday Call 2024-2025",
              "Prior Holiday Call 2025-2026",
              "NO CALL Holiday Request", "Call Holiday Preference",
              "NO CALL Weekend Request", "NO CALL Week Request",
              "Vac", "Acad", "Leave",
              "Prior Holiday Call ALL (Formatted)", "NO CALL Holiday Request (Formatted)",
              "NO CALL Weekend Request (Formatted)", "Vac/Acad/Leave (Formatted)",
              "NO NF ASSIGNMENTS", "Stack ok?", "Gad Preference", "Comment"]
ws_nc.append(nc_headers)

for r in residents:
    no_call_dates = []
    for _ in range(random.randint(2, 6)):
        month = random.randint(7, 12) if random.random() < 0.5 else random.randint(1, 6)
        day = random.randint(1, 28)
        no_call_dates.append(f"{month}/{day}")

    nf_dates_str = f"{r['full_name']}:{', '.join(no_call_dates)}"

    row = [
        r["first_name"],
        r["last_name"],
        r["pgy"],
        r["full_name"],
        f"R{r['r_year']}",
        "", "", "",
        random.choice(["Christmas", "Thanksgiving", "New Years"]),
        "",
        "",
        "",
        "10/15, 3/20",
        "11/10",
        "",
        "", "", "", "",
        nf_dates_str,
        random.choice(["Yes", "No"]),
        random.choice(["None", "Some", "A lot"]),
        "",
    ]
    ws_nc.append(row)

# === CallGad Responses (minimal) ===
ws_cg = wb.create_sheet("CallGad Responses")
cg_headers = ["First Name", "Last Name", "Current PGY", "Name", "R#"]
ws_cg.append(cg_headers)
for r in residents:
    ws_cg.append([r["first_name"], r["last_name"], r["pgy"], r["full_name"], f"R{r['r_year']}"])

# Save
output_path = "tests/dummy_prefs_2026_2027.xlsx"
wb.save(output_path)
print(f"Wrote dummy preferences to {output_path}")
print(f"  R1 Rotations: {len(r1s)} rows")
print(f"  R2 Rotations: {len(r2s)} rows")
print(f"  R3 Rotations: {len(r3s)} rows")
print(f"  R4 Rotations: {len(r4s)} rows")
print(f"  No Call Pref: {len(residents)} rows")
